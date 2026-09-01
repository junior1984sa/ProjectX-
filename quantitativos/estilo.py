# -*- coding: utf-8 -*-
"""Camada de apresentacao compartilhada pelas planilhas de quantitativo.

Cada prancha gera um arquivo .xlsx proprio, mas todos usam a mesma identidade
visual: paleta, fontes, cartoes de indicador, cabecalhos de tabela, blocos de
secao, graficos e configuracao de impressao. Este modulo concentra tudo isso
para que os arquivos saiam identicos em estilo, sem duplicar codigo.

Uso tipico em um script de prancha:

    from estilo import *
    wb = Workbook()
    ws = wb.active; ws.title = "Dashboard Executivo"
    painel_titulo(ws, "PAINEL ...", "Prancha ... | Escopo ...", ncols=14)
    card(ws, 1, 4, "MASSA TOTAL", 30598.9, "kg", AZUL, N_2)
    ...
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------- paleta
NAVY    = "FF13283F"   # faixa de titulo
CREME   = "FFF5F2EC"   # faixa de subtitulo
AZUL    = "FF195E83"   # cartao / serie 1
LARANJA = "FFD9652B"   # cartao / serie 2
ACO     = "FF31556F"   # cabecalhos de tabela e secoes
VERDE   = "FF5E8B76"   # cartao / serie 4
VALBG   = "FFFBFAF7"   # fundo do valor do cartao
TINTA   = "FF182532"   # texto corrente
LINHA   = "FFF8FAFB"   # fundo das linhas de tabela
BRANCO  = "FFFFFFFF"
ALERTA  = "FFFDF3E3"   # fundo das linhas de divergencia
ALERTA_TXT = "FF8A4B0F"
BORDA   = "FFB9C4CD"

CORES_SERIE = (AZUL, LARANJA, ACO, VERDE)

THIN = Side(style="thin", color=BORDA)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------- fontes
F_TIT   = Font(name="Arial", size=20, bold=True, color=BRANCO)
F_SUB   = Font(name="Arial", size=9, color=TINTA)
F_CARD  = Font(name="Arial", size=9, bold=True, color=BRANCO)
F_VAL   = Font(name="Arial", size=20, bold=True, color=TINTA)
F_CAP   = Font(name="Arial", size=8, color=TINTA)
F_SEC   = Font(name="Arial", size=11, bold=True, color=BRANCO)
F_HDR   = Font(name="Arial", size=10, bold=True, color=BRANCO)
F_BODY  = Font(name="Arial", size=9, color=TINTA)
F_BOLD  = Font(name="Arial", size=9, bold=True, color=TINTA)
F_SMALL = Font(name="Arial", size=8, color=TINTA)
F_TOTAL = Font(name="Arial", size=9, bold=True, color=BRANCO)

# ---------------------------------------------------------------- alinhamentos
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTRV = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP  = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---------------------------------------------------------------- formatos
N_INT = "#,##0"
N_2   = "#,##0.00"
N_3   = "#,##0.000"
N_PCT = "0.0%"


def _pintar(ws, linha, col_ini, col_fim, cor):
    """Aplica fundo e borda em um intervalo, inclusive nas celulas mescladas."""
    for j in range(col_ini, col_fim + 1):
        c = ws.cell(row=linha, column=j)
        c.fill = PatternFill("solid", fgColor=cor)
        c.border = BOX


def cabecalho(ws, titulo, subtitulo, ncols):
    """Faixa de titulo (navy) e de fonte (creme) no topo de uma aba."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = F_TIT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CTRV
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitulo)
    c.font = F_SUB
    c.fill = PatternFill("solid", fgColor=CREME)
    c.alignment = CTRV
    ws.row_dimensions[2].height = 24

    ws.sheet_view.showGridLines = False


def nota(ws, linha, ncols, texto, altura=26):
    """Linha de nota explicativa em corpo pequeno, mesclada na largura da aba."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = F_SMALL
    c.alignment = LEFT
    ws.row_dimensions[linha].height = altura


def secao(ws, linha, col_ini, col_fim, texto):
    """Barra de secao (cabecalho de bloco) dentro de uma aba."""
    ws.merge_cells(start_row=linha, start_column=col_ini, end_row=linha, end_column=col_fim)
    c = ws.cell(row=linha, column=col_ini, value=texto)
    c.font = F_SEC
    c.alignment = CTRV
    _pintar(ws, linha, col_ini, col_fim, ACO)
    c.font = F_SEC
    ws.row_dimensions[linha].height = 22


def tabela(ws, linha, headers, dados, formatos, larguras, fill=ACO, altura_hdr=30):
    """Escreve uma tabela com cabecalho colorido e devolve a ultima linha usada."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=linha, column=j, value=h)
        c.font = F_HDR
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CTR
        c.border = BOX
    ws.row_dimensions[linha].height = altura_hdr

    for i, row in enumerate(dados, start=linha + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = F_BODY
            c.fill = PatternFill("solid", fgColor=LINHA)
            c.border = BOX
            if isinstance(v, str):
                c.alignment = TOP if len(v) > 40 else LEFT
            else:
                c.alignment = CTRV
            if formatos[j - 1]:
                c.number_format = formatos[j - 1]

    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return linha + len(dados)


def total_row(ws, linha, ncols, rotulo, valores, formatos, span=1):
    """Linha de total em fundo solido; `span` colunas iniciais levam o rotulo."""
    if span > 1:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=span)
    c = ws.cell(row=linha, column=1, value=rotulo)
    c.font = F_TOTAL
    c.alignment = CTRV
    _pintar(ws, linha, 1, span, ACO)
    c.font = F_TOTAL

    for j, v in enumerate(valores, start=span + 1):
        c = ws.cell(row=linha, column=j, value=v)
        c.font = F_TOTAL
        c.fill = PatternFill("solid", fgColor=ACO)
        c.alignment = CTRV
        c.border = BOX
        if formatos[j - span - 1]:
            c.number_format = formatos[j - span - 1]


def card(ws, col, linha, titulo, valor, legenda, cor, fmt, largura=3):
    """Cartao de indicador do painel: titulo colorido, valor grande e legenda."""
    L = get_column_letter(col)
    R = get_column_letter(col + largura - 1)

    ws.merge_cells("%s%d:%s%d" % (L, linha, R, linha))
    c = ws.cell(row=linha, column=col, value=titulo)
    c.font = F_CARD
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = CTR

    ws.merge_cells("%s%d:%s%d" % (L, linha + 1, R, linha + 2))
    c = ws.cell(row=linha + 1, column=col, value=valor)
    c.font = F_VAL
    c.fill = PatternFill("solid", fgColor=VALBG)
    c.alignment = CTRV
    c.number_format = fmt

    ws.merge_cells("%s%d:%s%d" % (L, linha + 3, R, linha + 3))
    c = ws.cell(row=linha + 3, column=col, value=legenda)
    c.font = F_CAP
    c.alignment = CTRV

    for rr in (linha, linha + 1, linha + 2):
        for cc in range(col, col + largura):
            ws.cell(row=rr, column=cc).border = BOX


def _limpar_eixos(ch):
    ch.style = 2
    ch.y_axis.majorGridlines = None
    for ax in (ch.x_axis, ch.y_axis):
        ax.txPr = None


def grafico_barras(ws_dados, col_valores, lin_hdr, lin_fim, col_rotulos,
                   titulo, eixo_y, eixo_x, cor=AZUL, largura=15.6, altura=8.4):
    """Grafico de colunas. `lin_hdr` e a linha do cabecalho (vira o nome da serie)."""
    ch = BarChart()
    ch.type = "col"
    ch.title = titulo
    ch.add_data(Reference(ws_dados, min_col=col_valores, min_row=lin_hdr, max_row=lin_fim),
                titles_from_data=True)
    ch.set_categories(Reference(ws_dados, min_col=col_rotulos, min_row=lin_hdr + 1, max_row=lin_fim))
    ch.y_axis.title = eixo_y
    ch.x_axis.title = eixo_x
    ch.height = altura
    ch.width = largura
    ch.legend = None
    ch.series[0].graphicalProperties = GraphicalProperties(solidFill=cor[2:])
    _limpar_eixos(ch)
    return ch


def grafico_rosca(ws_dados, col_valores, lin_hdr, lin_fim, col_rotulos,
                  titulo, cores=CORES_SERIE, largura=15.6, altura=8.4):
    """Grafico de rosca com uma cor explicita por fatia."""
    ch = DoughnutChart(holeSize=55)
    ch.title = titulo
    ch.add_data(Reference(ws_dados, min_col=col_valores, min_row=lin_hdr, max_row=lin_fim),
                titles_from_data=True)
    ch.set_categories(Reference(ws_dados, min_col=col_rotulos, min_row=lin_hdr + 1, max_row=lin_fim))
    ch.height = altura
    ch.width = largura
    for idx in range(lin_fim - lin_hdr):
        cor = cores[idx % len(cores)]
        ch.series[0].data_points.append(
            DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=cor[2:])))
    return ch


def configurar_impressao(wb, aba_pagina_unica="Dashboard Executivo", titulos_repetidos="1:5"):
    """Paisagem A4, ajuste a largura; o painel cabe em uma unica pagina."""
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1 if ws.title == aba_pagina_unica else 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4
        if ws.title != aba_pagina_unica and titulos_repetidos:
            ws.print_title_rows = titulos_repetidos
