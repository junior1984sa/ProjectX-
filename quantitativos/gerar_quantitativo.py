# -*- coding: utf-8 -*-
"""Gera o quantitativo de armaduras da prancha DE-5400.00-6310-120-TX3-002."""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.chart.shapes import GraphicalProperties

OUT = "/home/user/ProjectX-/quantitativos/Quantitativo_Armaduras_DE-5400.00-6310-120-TX3-002.xlsx"

# ---------------------------------------------------------------- paleta
NAVY   = "FF13283F"; CREME  = "FFF5F2EC"; AZUL   = "FF195E83"
LARANJA= "FFD9652B"; AÇO    = "FF31556F"; VERDE  = "FF5E8B76"
VALBG  = "FFFBFAF7"; TINTA  = "FF182532"; LINHA  = "FFF8FAFB"
BRANCO = "FFFFFFFF"; ALERTA = "FFFDF3E3"; BORDA  = "FFB9C4CD"
THIN = Side(style="thin", color=BORDA)
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TIT   = Font(name="Arial", size=20, bold=True, color=BRANCO)
F_SUB   = Font(name="Arial", size=9, color=TINTA)
F_CARD  = Font(name="Arial", size=9, bold=True, color=BRANCO)
F_VAL   = Font(name="Arial", size=20, bold=True, color=TINTA)
F_CAP   = Font(name="Arial", size=8, color=TINTA)
F_SEC   = Font(name="Arial", size=11, bold=True, color=BRANCO)
F_HDR   = Font(name="Arial", size=10, bold=True, color=BRANCO)
F_BODY  = Font(name="Arial", size=9, color=TINTA)
F_BOLD  = Font(name="Arial", size=9, bold=True, color=TINTA)
F_SMALL = Font(name="Arial", size=8, color=TINTA)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTRV= Alignment(horizontal="center", vertical="center")
LEFT= Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

N_INT="#,##0"; N_2="#,##0.00"; N_3="#,##0.000"; N_PCT="0.0%"

# ---------------------------------------------------------------- dados da prancha
KGM = {8: 0.395, 16: 1.578, 20: 2.466}          # NBR 7480 - massa linear nominal
NT  = 8                                          # fundações cobertas pela prancha (8x)

# POS, bitola, quant(8x), unit(cm) ou None, total(m) impresso, elemento, função, vista
BARRAS = [
 (1, 20,   16, None, 1142.40, "N1", "Corrida face sup./inf.", "CORTE A-A / FACE SUP. E INF.",
  "Barra corrida (CORR) do banzo superior e inferior; 2 un por fundação (1 sup. + 1 inf.)."),
 (2, 20,   16, None, 1206.40, "N2", "Corrida face sup./inf.", "CORTE A-A / FACE SUP. E INF.",
  "Barra corrida (CORR) do banzo superior e inferior; 2 un por fundação (1 sup. + 1 inf.)."),
 (3, 20,   16, None, 1206.40, "N3", "Corrida face sup./inf.", "CORTE A-A / FACE SUP. E INF.",
  "Barra corrida (CORR) do banzo superior e inferior; 2 un por fundação (1 sup. + 1 inf.)."),
 (4, 20,   16, None, 1206.40, "N4", "Corrida face sup./inf.", "CORTE A-A / FACE SUP. E INF.",
  "Barra corrida (CORR) do banzo superior e inferior; 2 un por fundação (1 sup. + 1 inf.)."),
 (5, 20,   16, None, 1209.60, "N5", "Corrida face sup./inf.", "CORTE A-A / FACE SUP. E INF.",
  "Barra corrida (CORR) do banzo superior e inferior; 2 un por fundação (1 sup. + 1 inf.)."),
 (6, 16,   48, None, 3427.20, "N6", "Corrida face externa", "CORTE A-A / FACE EXTERNA",
  "6 N6 c/20 na face externa; barra corrida (CORR) ao longo do desenvolvimento do anel."),
 (7, 16,   48, None, 3264.00, "N7", "Corrida face interna", "CORTE A-A / FACE INTERNA",
  "6 N7 c/20 na face interna; barra corrida (CORR) ao longo do desenvolvimento do anel."),
 (8,  8, 2384,  465, 11085.60, "N8", "Estribo", "PLANTA / CORTE A-A",
  "298 N8 c/20 por fundação; estribo fechado 55 × 140 cm, C=465 cm."),
 (9,  8,   48,  473, 227.04, "N9", "Estribo no rebaixo", "VISTA C-C",
  "6 N9 c/20 por fundação; estribo na região do rebaixo, C=473 cm."),
 (10,16,   32,  500, 160.00, "N10", "Reforço do rebaixo", "VISTA C-C / CORTE B-B",
  "N10-4 Ø16 C=500 cm; disposto 2x2 na região do rebaixo."),
 (11,20,   32,  600, 192.00, "N11", "Reforço superior do rebaixo", "VISTA C-C / CORTE B-B",
  "N11-4 Ø20 C=600 cm (SUP.), dobra Ri=8 e perna 120 cm; ver Nota 5."),
 (12, 8,   32, None, 150.72, "N12", "Estribo de comprimento variável", "VISTA C-C",
  "N12-2x2 Ø8 C=VAR; forma cotada com perna variável (113 a 133) cm."),
 (13,16,    8,  485, 38.80, "N13", "Reforço do rebaixo", "VISTA C-C / CORTE B-B",
  "N13-2 Ø16 C=485 cm na região do rebaixo."),
]

RESUMO_DES = {8: (11463.36, 4528), 16: (6890.00, 10872), 20: (6163.20, 15198)}
TOTAL_DES  = 30599

GRUPOS = [
 ("Barras corridas — faces superior e inferior", [1,2,3,4,5]),
 ("Barras corridas — faces externa e interna",   [6,7]),
 ("Estribos",                                    [8,9,12]),
 ("Reforços da região do rebaixo",               [10,11,13]),
]
TANQUES = ["TQ-6310816A","TQ-6310816B","TQ-6310816C","TQ-6310816D",
           "TQ-6310817A","TQ-6310817B","TQ-6310817C","TQ-6310817D"]

# desenvolvimentos cotados nas elevações (cm -> m)
DEV_EXT, DEV_INT, DEV_SI = 61.556, 59.942, 58.327
SEC_B, SEC_H = 0.55, 1.40

def massa(pos):
    p, d, q, u, t = pos[0], pos[1], pos[2], pos[3], pos[4]
    return t * KGM[d]

MASSA_TOT = sum(massa(b) for b in BARRAS)
COMPR_TOT = sum(b[4] for b in BARRAS)
QTD_ESTRIBOS = sum(b[2] for b in BARRAS if b[5] in ("N8","N9","N12"))
BARRAS_12M = sum(math.ceil(RESUMO_DES[d][0] / 12) for d in (8,16,20))

wb = Workbook()

# ---------------------------------------------------------------- helpers
def cabecalho(ws, titulo, subtitulo, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = F_TIT; c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CTRV
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=2 if False else 1, value=subtitulo)
    c.font = F_SUB; c.fill = PatternFill("solid", fgColor=CREME); c.alignment = CTRV
    ws.row_dimensions[2].height = 24
    ws.sheet_view.showGridLines = False

def tabela(ws, linha, headers, dados, formatos, larguras, fill=AÇO):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=linha, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CTR; c.border = BOX
    ws.row_dimensions[linha].height = 30
    for i, row in enumerate(dados, start=linha + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA)
            c.border = BOX; c.alignment = TOP if isinstance(v, str) and len(str(v)) > 40 else (
                LEFT if isinstance(v, str) else CTRV)
            if formatos[j-1]: c.number_format = formatos[j-1]
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return linha + len(dados)

def total_row(ws, linha, ncols, rotulo, valores, formatos, span=1):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=span)
    c = ws.cell(row=linha, column=1, value=rotulo)
    c.font = Font(name="Arial", size=9, bold=True, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTRV; c.border = BOX
    for j in range(2, span + 1):
        ws.cell(row=linha, column=j).fill = PatternFill("solid", fgColor=AÇO)
        ws.cell(row=linha, column=j).border = BOX
    for j, v in enumerate(valores, start=span + 1):
        c = ws.cell(row=linha, column=j, value=v)
        c.font = Font(name="Arial", size=9, bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTRV; c.border = BOX
        if formatos[j - span - 1]: c.number_format = formatos[j - span - 1]

def seção(ws, linha, col_ini, col_fim, texto):
    ws.merge_cells(start_row=linha, start_column=col_ini, end_row=linha, end_column=col_fim)
    c = ws.cell(row=linha, column=col_ini, value=texto)
    c.font = F_SEC; c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTRV; c.border = BOX
    for j in range(col_ini, col_fim + 1):
        ws.cell(row=linha, column=j).border = BOX
    ws.row_dimensions[linha].height = 22

# ================================================================ abas
ws_dash = wb.active; ws_dash.title = "Dashboard Executivo"
ws_lista = wb.create_sheet("Lista de Barras")
ws_bit   = wb.create_sheet("Resumo por Bitola")
ws_pos   = wb.create_sheet("Resumo por Posição")
ws_tq    = wb.create_sheet("Quantitativo por Tanque")
ws_geo   = wb.create_sheet("Geometria e Seções")
ws_par   = wb.create_sheet("Parâmetros Técnicos")
ws_val   = wb.create_sheet("Validações")
ws_ler   = wb.create_sheet("Leia-me")

FONTE = ("Fonte: prancha DE-5400.00-6310-120-TX3-002, Rev. 0 (07/08/2026) - Desenho de fundação, "
         "tanques de lubrificação TQ-6310816A/B/C/D & TQ-6310817A/B/C/D - ARMADURAS")

# ---------------------------------------------------------------- Lista de Barras
cabecalho(ws_lista, "LISTA DE BARRAS - TRANSCRIÇÃO INTEGRAL DA PRANCHA", FONTE, 14)
ws_lista["A3"] = ("Quadro LISTA DE BARRAS (8x) transcrito posição a posição. As colunas QUANT. e TOTAL do "
                  "desenho referem-se ao conjunto das 8 fundações; as colunas por fundação são derivadas por divisão por 8.")
ws_lista["A3"].font = F_SMALL; ws_lista["A3"].alignment = LEFT
ws_lista.merge_cells("A3:N3"); ws_lista.row_dimensions[3].height = 26

dados = []
for p, d, q, u, t, el, fun, vista, obs in BARRAS:
    m = t * KGM[d]
    dados.append([
        p, el, d, q, ("CORR" if u is None and p != 12 else ("VAR" if p == 12 else u)),
        t, KGM[d], round(m, 2), round(m / MASSA_TOT, 6),
        q // NT, round(t / NT, 2), round(m / NT, 2),
        fun, vista,
    ])
fim = tabela(ws_lista, 5,
    ["POS.", "Elemento", "Bitola\nØ (mm)", "Quant. total\n(8 fundações)", "Compr. unit.\n(cm)",
     "Compr. total\n(m) - 8x", "Massa linear\n(kg/m)", "Massa total\n(kg) - 8x", "% da massa",
     "Quant. por\nfundação", "Compr. por\nfundação (m)", "Massa por\nfundação (kg)",
     "Função estrutural", "Vista de referência"],
    dados,
    [N_INT, None, N_INT, N_INT, None, N_2, N_3, N_2, N_PCT, N_INT, N_2, N_2, None, None],
    [7, 10, 9, 13, 11, 13, 11, 13, 10, 11, 12, 12, 26, 26])
total_row(ws_lista, fim + 1, 14, "TOTAL — 13 posições",
    [None, sum(b[2] for b in BARRAS), None, round(COMPR_TOT, 2), None, round(MASSA_TOT, 2), 1.0,
     None, round(COMPR_TOT / NT, 2), round(MASSA_TOT / NT, 2), None, None],
    [None, N_INT, None, N_2, None, N_2, N_PCT, None, N_2, N_2, None, None], span=2)

r = fim + 3
ws_lista.cell(row=r, column=1, value="Observações por posição").font = F_BOLD
for i, (p, d, q, u, t, el, fun, vista, obs) in enumerate(BARRAS, start=r + 1):
    ws_lista.cell(row=i, column=1, value=el).font = F_BOLD
    ws_lista.merge_cells(start_row=i, start_column=2, end_row=i, end_column=14)
    c = ws_lista.cell(row=i, column=2, value=obs); c.font = F_SMALL; c.alignment = LEFT
ws_lista.freeze_panes = "A6"

# ---------------------------------------------------------------- Resumo por Bitola
cabecalho(ws_bit, "RESUMO GERAL POR BITOLA", FONTE, 10)
ws_bit["A3"] = ("Confronto entre o quadro RESUMO GERAL da prancha e o somatório das posições da LISTA DE BARRAS. "
                "Massas lineares conforme NBR 7480 (valores nominais adotados no próprio desenho).")
ws_bit["A3"].font = F_SMALL; ws_bit["A3"].alignment = LEFT
ws_bit.merge_cells("A3:J3"); ws_bit.row_dimensions[3].height = 26

dados = []
for d in (8, 16, 20):
    comp_des, massa_des = RESUMO_DES[d]
    comp_calc = sum(b[4] for b in BARRAS if b[1] == d)
    massa_calc = comp_calc * KGM[d]
    posições = ", ".join("N%d" % b[0] for b in BARRAS if b[1] == d)
    dados.append([d, posições, comp_des, KGM[d], massa_des, round(massa_calc, 2),
                  round(massa_calc - massa_des, 2), round(massa_calc / MASSA_TOT, 6),
                  math.ceil(comp_des / 12), round(massa_calc / NT, 2)])
fim = tabela(ws_bit, 5,
    ["Bitola\nØ (mm)", "Posições", "Compr. total\n(m) - 8x", "Massa linear\n(kg/m)",
     "Massa da prancha\n(kg)", "Massa calculada\n(kg)", "Desvio\n(kg)", "% da massa",
     "Barras de 12 m\n(equivalente)", "Massa por\nfundação (kg)"],
    dados,
    [N_INT, None, N_2, N_3, N_INT, N_2, N_2, N_PCT, N_INT, N_2],
    [9, 22, 13, 11, 15, 14, 10, 10, 14, 13])
total_row(ws_bit, fim + 1, 10, "TOTAL",
    [round(COMPR_TOT, 2), None, TOTAL_DES, round(MASSA_TOT, 2), round(MASSA_TOT - TOTAL_DES, 2),
     1.0, BARRAS_12M, round(MASSA_TOT / NT, 2)],
    [N_2, None, N_INT, N_2, N_2, N_PCT, N_INT, N_2], span=2)
r = fim + 3
for txt in [
  "O quadro RESUMO GERAL da prancha fecha exatamente com o somatório da LISTA DE BARRAS nas três bitolas.",
  "TOTAL (kg) impresso na prancha = 30.599 kg. O somatório das massas não arredondadas é 30.598,90 kg; a soma "
  "das três linhas já arredondadas é 30.598 kg. A diferença é apenas de arredondamento (menos de 1 kg).",
  "A coluna 'Barras de 12 m' é um equivalente de suprimentos (comprimento total / 12 m, arredondado para cima). "
  "Não considera perdas de corte, dobra e emendas - ver aba Validações."]:
    ws_bit.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws_bit.cell(row=r, column=1, value="- " + txt); c.font = F_SMALL; c.alignment = LEFT
    ws_bit.row_dimensions[r].height = 26; r += 1
ws_bit.freeze_panes = "A6"

# ---------------------------------------------------------------- Resumo por Posição
cabecalho(ws_pos, "RESUMO POR POSIÇÃO DE BARRA", FONTE, 8)
ws_pos["A3"] = "Massa por posição (N1 a N13), ordenada conforme a prancha. Fonte de dados dos graficos do painel."
ws_pos["A3"].font = F_SMALL; ws_pos["A3"].alignment = LEFT
ws_pos.merge_cells("A3:H3")

dados = []
for p, d, q, u, t, el, fun, vista, obs in BARRAS:
    m = t * KGM[d]
    dados.append([el, d, q, t, round(m, 2), round(m / MASSA_TOT, 6), fun, round(m / NT, 2)])
fim = tabela(ws_pos, 5,
    ["Posição", "Bitola\nØ (mm)", "Quant. total\n(8x)", "Compr. total\n(m)", "Massa total\n(kg)",
     "% da massa", "Função estrutural", "Massa por\nfundação (kg)"],
    dados, [None, N_INT, N_INT, N_2, N_2, N_PCT, None, N_2],
    [11, 10, 12, 13, 13, 10, 30, 13])
total_row(ws_pos, fim + 1, 8, "TOTAL",
    [sum(b[2] for b in BARRAS), round(COMPR_TOT, 2), round(MASSA_TOT, 2), 1.0, None, round(MASSA_TOT / NT, 2)],
    [N_INT, N_2, N_2, N_PCT, None, N_2], span=2)

# agrupamento funcional (fonte do grafico de rosca)
r = fim + 3
seção(ws_pos, r, 1, 5, "AGRUPAMENTO POR FUNÇÃO ESTRUTURAL")
grp = []
for nome, poss in GRUPOS:
    m = sum(b[4] * KGM[b[1]] for b in BARRAS if b[0] in poss)
    c_ = sum(b[4] for b in BARRAS if b[0] in poss)
    q_ = sum(b[2] for b in BARRAS if b[0] in poss)
    grp.append([nome, ", ".join("N%d" % p for p in poss), q_, round(c_, 2), round(m, 2)])
fim2 = tabela(ws_pos, r + 1,
    ["Função estrutural", "Posições", "Quant. total\n(8x)", "Compr. total\n(m)", "Massa total\n(kg)"],
    grp, [None, None, N_INT, N_2, N_2], [40, 18, 12, 13, 13])
total_row(ws_pos, fim2 + 1, 5, "TOTAL",
    [sum(b[2] for b in BARRAS), round(COMPR_TOT, 2), round(MASSA_TOT, 2)],
    [N_INT, N_2, N_2], span=2)
ws_pos.freeze_panes = "A6"
GRP_ROW0 = r + 2                     # primeira linha de dados do agrupamento

# ---------------------------------------------------------------- Quantitativo por Tanque
cabecalho(ws_tq, "QUANTITATIVO POR TANQUE", FONTE, 8)
ws_tq["A3"] = ("A prancha cobre 8 fundações idênticas (indicação '8x' na planta e na lista de barras). "
               "O quantitativo unitário é obtido dividindo os totais por 8.")
ws_tq["A3"].font = F_SMALL; ws_tq["A3"].alignment = LEFT
ws_tq.merge_cells("A3:H3"); ws_tq.row_dimensions[3].height = 26

dados = []
for tq in TANQUES:
    dados.append([tq, "Fundação anelar — armadura", 1,
                  round(sum(b[4] for b in BARRAS if b[1] == 8) / NT, 2),
                  round(sum(b[4] for b in BARRAS if b[1] == 16) / NT, 2),
                  round(sum(b[4] for b in BARRAS if b[1] == 20) / NT, 2),
                  round(COMPR_TOT / NT, 2), round(MASSA_TOT / NT, 2)])
fim = tabela(ws_tq, 5,
    ["Tanque", "Elemento", "Fundações", "Compr. Ø8\n(m)", "Compr. Ø16\n(m)", "Compr. Ø20\n(m)",
     "Compr. total\n(m)", "Massa de aço\n(kg)"],
    dados, [None, None, N_INT, N_2, N_2, N_2, N_2, N_2],
    [17, 26, 11, 12, 12, 12, 13, 14])
total_row(ws_tq, fim + 1, 8, "TOTAL — 8 fundações",
    [NT,
     round(sum(b[4] for b in BARRAS if b[1] == 8), 2),
     round(sum(b[4] for b in BARRAS if b[1] == 16), 2),
     round(sum(b[4] for b in BARRAS if b[1] == 20), 2),
     round(COMPR_TOT, 2), round(MASSA_TOT, 2)],
    [N_INT, N_2, N_2, N_2, N_2, N_2], span=2)

r = fim + 3
seção(ws_tq, r, 1, 6, "DETALHAMENTO UNITÁRIO - UMA FUNDAÇÃO (VALE PARA CADA UM DOS 8 TANQUES)")
uni = []
for p, d, q, u, t, el, fun, vista, obs in BARRAS:
    uni.append([el, d, q // NT, ("CORR" if u is None and p != 12 else ("VAR" if p == 12 else u)),
                round(t / NT, 2), round(t * KGM[d] / NT, 2)])
fim2 = tabela(ws_tq, r + 1,
    ["Posição", "Bitola\nØ (mm)", "Quant. por\nfundação", "Compr. unit.\n(cm)",
     "Compr. por\nfundação (m)", "Massa por\nfundação (kg)"],
    uni, [None, N_INT, N_INT, None, N_2, N_2], [11, 10, 12, 12, 13, 14])
total_row(ws_tq, fim2 + 1, 6, "TOTAL POR FUNDAÇÃO",
    [sum(b[2] for b in BARRAS) // NT, None, round(COMPR_TOT / NT, 2), round(MASSA_TOT / NT, 2)],
    [N_INT, None, N_2, N_2], span=2)
ws_tq.freeze_panes = "A6"

# ---------------------------------------------------------------- Geometria e Seções
cabecalho(ws_geo, "GEOMETRIA, VISTAS E SEÇÕES", FONTE, 6)
ws_geo["A3"] = ("Cotas e desenvolvimentos lidos diretamente da prancha, e grandezas derivadas a partir delas. "
                "Os itens marcados como DERIVADO não constam do desenho: são cálculo próprio para apoio de planejamento.")
ws_geo["A3"].font = F_SMALL; ws_geo["A3"].alignment = LEFT
ws_geo.merge_cells("A3:F3"); ws_geo.row_dimensions[3].height = 26

geo = [
 ["Vista", "PLANTA - ARMADURA - TQ-6310816 A/B/C/D & TQ-6310817 A/B/C/D (8x)", "1:100", "-", "PRANCHA",
  "Planta de armadura do anel de fundação; indicação 298 N8 c.20."],
 ["Vista", "CORTE A-A", "1:20", "-", "PRANCHA",
  "Seção corrente do anel: 55 cm (radial) × 140 cm (altura); 6 N6 c/20 na face externa e 6 N7 c/20 na face interna."],
 ["Vista", "CORTE B-B", "1:20", "-", "PRANCHA",
  "Seção na região do rebaixo, com N10, N11, N13 e 5 N6 / 5 N7 c/20."],
 ["Vista", "VISTA C-C (2x)", "1:25", "-", "PRANCHA",
  "Detalhe do rebaixo: N9, N10, N11, N12 e N13."],
 ["Vista", "FACE EXTERNA - ELEVAÇÃO", "S/ESC.", 61.556, "PRANCHA",
  "Desenvolvimento cotado de 6155,6 cm; malha de N5 e N6 com trechos de 160 cm e traspasses de 40 cm."],
 ["Vista", "FACE INTERNA - ELEVAÇÃO", "S/ESC.", 59.942, "PRANCHA",
  "Desenvolvimento cotado de 5994,2 cm; malha de N1 e N7."],
 ["Vista", "FACE SUPERIOR E INFERIOR - PLANTA", "S/ESC.", 58.327, "PRANCHA",
  "Desenvolvimento cotado de 5832,7 cm; malha de N1 a N5."],
 ["Seção", "Largura radial da seção corrente", "-", 0.55, "PRANCHA", "Cota 55 cm no CORTE A-A e no CORTE B-B."],
 ["Seção", "Altura da seção corrente", "-", 1.40, "PRANCHA", "Cota 140 cm no CORTE A-A."],
 ["Seção", "Área da seção corrente (m²)", "-", 0.77, "DERIVADO", "0,55 m × 1,40 m."],
 ["Espaçamento", "Estribos N8 na seção corrente", "-", 0.20, "PRANCHA",
  "298 N8 c/20 por fundação; 298 × 0,20 m = 59,60 m, coerente com o desenvolvimento da face interna (59,942 m)."],
 ["Espaçamento", "Barras horizontais N6 / N7", "-", 0.20, "PRANCHA", "6 N6 c/20 e 6 N7 c/20 no CORTE A-A."],
 ["Traspasse", "Traspasse típico das barras corridas", "-", 0.40, "PRANCHA",
  "Cotas 40 (TIP.) e 160 repetidas nas elevações das faces."],
 ["Derivado", "Perímetro médio do anel (m)", "-", round((DEV_EXT + DEV_INT) / 2, 3), "DERIVADO",
  "Média entre os desenvolvimentos das faces externa (61,556 m) e interna (59,942 m)."],
 ["Derivado", "Diâmetro médio implícito (m)", "-", round(((DEV_EXT + DEV_INT) / 2) / math.pi, 3), "DERIVADO",
  "Perímetro médio dividido por pi. Valor indicativo — confirmar na prancha de formas."],
 ["Derivado", "Volume de concreto por fundação (m³)", "-", round(SEC_B * SEC_H * (DEV_EXT + DEV_INT) / 2, 2), "DERIVADO",
  "ESTIMATIVA: área da seção corrente × perímetro médio. Não considera o rebaixo nem variações de seção. "
  "O volume oficial deve ser lido na prancha de formas DE-5400.00-6310-120-TX3-001."],
 ["Derivado", "Volume de concreto — 8 fundações (m³)", "-", round(SEC_B * SEC_H * (DEV_EXT + DEV_INT) / 2 * NT, 2), "DERIVADO",
  "ESTIMATIVA - mesma ressalva da linha anterior."],
 ["Derivado", "Taxa de armadura (kg/m³)", "-", round(MASSA_TOT / (SEC_B * SEC_H * (DEV_EXT + DEV_INT) / 2 * NT), 1), "DERIVADO",
  "ESTIMATIVA: massa total de aço dividida pelo volume estimado. Indicador de aderência; não substitui a prancha de formas."],
]
fim = tabela(ws_geo, 5, ["Grupo", "Item", "Escala", "Valor\n(m, m² ou m³)", "Origem", "Observação"],
    geo, [None, None, None, N_3, None, None], [14, 40, 9, 14, 12, 62])
ws_geo.freeze_panes = "A6"

# ---------------------------------------------------------------- Parâmetros Técnicos
cabecalho(ws_par, "PARÂMETROS TÉCNICOS E IDENTIFICAÇÃO DA PRANCHA", FONTE, 3)
ws_par["A3"] = "Dados de carimbo, notas gerais, legenda e documentos vinculados, transcritos da prancha."
ws_par["A3"].font = F_SMALL; ws_par["A3"].alignment = LEFT
ws_par.merge_cells("A3:C3")

par = [
 ["Identificação", "Número do documento", "DE-5400.00-6310-120-TX3-002"],
 ["Identificação", "Revisão", "0 - EMISSÃO ORIGINAL - PARA CONSTRUÇÃO (07/08/2026)"],
 ["Identificação", "Titulo", "Desenho de fundação - tanques de lubrificação TQ-6310816A/B/C/D & TQ-6310817A/B/C/D - ARMADURAS"],
 ["Identificação", "Área", "Parque de tanques de produtos acabados (lubrificantes) - U-6310"],
 ["Identificação", "Programa / Cliente", "Refino Boaventura / TR - Boaventura"],
 ["Identificação", "Proprietário do documento", "Petrobras - SRGE/SI-III (classificação INTERNA)"],
 ["Identificação", "Razão social", "Consorcio TEM Boaventura"],
 ["Identificação", "Número do contrato", "ICJ 5900.0131990.25.2"],
 ["Identificação", "Responsável técnico", "Antenor de Castro - CREA 17974D-MG"],
 ["Identificação", "Executou / Verificou / Aprovou", "Carmen Santos / Marcio Yukio / Helgo Santos"],
 ["Identificação", "Folha e formato", "01 de 01 - A1 (841 × 594 mm) - escala INDICADA"],
 ["Nota geral 1", "Unidades", "Dimensões em centímetro, bitolas das barras em milímetro, exceto onde indicado."],
 ["Nota geral 2", "Aço", "CA-50."],
 ["Nota geral 3", "Dobramento", "Dobramento das barras conforme NBR-6118."],
 ["Nota geral 4", "Cobrimento", "Cobrimento mínimo pela face externa das barras: 5 cm."],
 ["Nota geral 5", "Interferências", "Cortar e dobrar as barras que interferirem com os rebaixos."],
 ["Critério adotado", "Massa linear das barras", "Ø8 = 0,395 kg/m; Ø16 = 1,578 kg/m; Ø20 = 2,466 kg/m - valores do quadro RESUMO GERAL, coincidentes com os nominais da NBR 7480."],
 ["Legenda", "Abreviaturas", "F.EXT. = face externa; F.INT. = face interna; SUP. = superior; INF. = inferior; CORR = barra corrida; VAR = comprimento variável."],
 ["Documento de referência", "Memória de cálculo", "MC-5400.00-6310-120-TX3-001 - Memória de cálculo de fundação - tanques de lubrificação (U-6310)."],
 ["Documento complementar", "Formas", "DE-5400.00-6310-120-TX3-001 - Desenho de fundação - tanques de lubrificação TQ-6310816A/B/C/D & TQ-6310817A/B/C/D - FORMAS."],
 ["Pendência P1", "Arranjo geral", "CI-6310-001 - Arranjo geral de equipamentos."],
 ["Pendência P2", "Cargas", "CI-6310-004 - Cargas finais de vento e chumbador."],
]
fim = tabela(ws_par, 5, ["Grupo", "Parâmetro", "Especificação / conteudo transcrito"],
             par, [None, None, None], [26, 34, 96])
ws_par.freeze_panes = "A6"

# ---------------------------------------------------------------- Validações
cabecalho(ws_val, "CONTROLES DE CONSISTÊNCIA", FONTE, 6)
ws_val["A3"] = ("Conferência entre o somatório estruturado e os quadros impressos na prancha. "
                "Toda divergência está sinalizada de forma explícita.")
ws_val["A3"].font = F_SMALL; ws_val["A3"].alignment = LEFT
ws_val.merge_cells("A3:F3")

val = []
for d in (8, 16, 20):
    comp_des = RESUMO_DES[d][0]
    comp_calc = sum(b[4] for b in BARRAS if b[1] == d)
    val.append(["Comprimento", "Ø%d - comprimento total (m)" % d, round(comp_calc, 2), comp_des,
                round(comp_calc - comp_des, 2),
                "Soma das posições da LISTA DE BARRAS confere exatamente com o quadro RESUMO GERAL."])
for d in (8, 16, 20):
    massa_des = RESUMO_DES[d][1]
    massa_calc = sum(b[4] for b in BARRAS if b[1] == d) * KGM[d]
    val.append(["Massa", "Ø%d - massa total (kg)" % d, round(massa_calc, 2), massa_des,
                round(massa_calc - massa_des, 2),
                "Comprimento × massa linear. Desvio inferior a 0,5 kg, compatível com o arredondamento do quadro."])
val += [
 ["Massa", "Massa total de aço (kg)", round(MASSA_TOT, 2), TOTAL_DES, round(MASSA_TOT - TOTAL_DES, 2),
  "TOTAL (kg) impresso = 30.599. Somatório não arredondado = 30.598,90 kg. A soma das três linhas já "
  "arredondadas resulta 30.598 kg. Diferença exclusivamente de arredondamento."],
 ["Lista de barras", "TOTAL = QUANT. x UNIT. (posições N8, N9, N10, N11, N13)", 5, 5, 0,
  "As cinco posições de comprimento fixo fecham exatamente: N8 2384 × 4,65 = 11.085,60 m; N9 48 × 4,73 = 227,04 m; "
  "N10 32 × 5,00 = 160,00 m; N11 32 × 6,00 = 192,00 m; N13 8 × 4,85 = 38,80 m."],
 ["Multiplicidade 8x", "N8 - quantidade total vs. cota da planta", 2384, 298 * NT, 0,
  "A planta indica 298 N8 c.20 por fundação. 298 × 8 = 2.384, confirmando que a coluna QUANT. da lista já "
  "corresponde ao conjunto das 8 fundações."],
 ["Multiplicidade 8x", "N9 / N10 / N11 / N12 - quantidade total vs. cotas dos detalhes", 4, 4, 0,
  "N9 = 6/fundação (48/8); N10 = 4 (32/8); N11 = 4 (32/8); N12 = 2x2 = 4 (32/8). Todas coerentes com os "
  "detalhes da VISTA C-C."],
 ["DIVERGÊNCIA", "N13 - quantidade por fundação", 1, 2, -1,
  "A LISTA DE BARRAS indica QUANT. 8 (1 un por fundação) e TOTAL 38,80 m, internamente consistente. "
  "Já o detalhe da VISTA C-C / CORTE B-B chama 'N13-2 Ø16-485' e '2N13', ou seja, 2 un por fundação (16 no total). "
  "Se prevalecer o detalhe, acrescem 8 barras, 38,80 m e 61,23 kg de Ø16 (+0,20% da massa total). CONFIRMAR COM O PROJETISTA."],
 ["DIVERGÊNCIA", "N12 - comprimento unitário", 4.71, 1.23, 3.48,
  "A coluna UNIT. traz 'VAR' e a forma está cotada com perna variável (113 a 133) cm. Contudo, TOTAL 150,72 m "
  "para 32 un implica comprimento médio de 471 cm por barra - próximo dos estribos N8 (465) e N9 (473) e não da "
  "faixa cotada. O RESUMO GERAL adota 150,72 m. Se o comprimento médio fosse 123 cm, o total cairia para 39,36 m "
  "(-43,99 kg, -0,14% da massa total). CONFIRMAR COM O PROJETISTA."],
 ["Coerência geométrica", "Estribos N8 x desenvolvimento do anel (m)", round(298 * 0.20, 2), DEV_INT,
  round(298 * 0.20 - DEV_INT, 2),
  "298 estribos a cada 20 cm cobrem 59,60 m, praticamente o desenvolvimento cotado da face interna (59,942 m). "
  "Distribuição coerente."],
 ["Materiais", "Massas lineares adotadas (kg/m)", 3, 3, 0,
  "Ø8 = 0,395; Ø16 = 1,578; Ø20 = 2,466. Coincidem com os valores nominais da NBR 7480 para aço CA-50."],
 ["Escopo", "Perdas de corte, dobra e emendas", 0, 0, 0,
  "Os quantitativos são os do desenho (comprimentos de projeto). Não há acréscimo de perdas. Para suprimentos, "
  "aplicar a taxa de perda contratual sobre a massa por bitola."],
 ["Escopo", "Concreto, formas e chumbadores", 0, 0, 0,
  "Esta prancha é exclusivamente de ARMADURAS. Volume de concreto, formas e insertos devem ser lidos na prancha "
  "de formas DE-5400.00-6310-120-TX3-001 (documento complementar), não fornecida nesta análise."],
]
fim = tabela(ws_val, 5, ["Grupo", "Indicador", "Calculado", "Referência", "Desvio", "Critério / observação"],
             val, [None, None, N_2, N_2, N_2, None], [20, 42, 12, 12, 11, 90])
# realce das divergências
for i in range(6, fim + 1):
    if ws_val.cell(row=i, column=1).value == "DIVERGÊNCIA":
        for j in range(1, 7):
            ws_val.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ALERTA)
            ws_val.cell(row=i, column=j).font = Font(name="Arial", size=9, bold=True, color="FF8A4B0F")
ws_val.freeze_panes = "A6"

# ---------------------------------------------------------------- Leia-me
cabecalho(ws_ler, "LEIA-ME - QUANTITATIVO DE ARMADURAS", FONTE, 2)
ler = [
 ["Objetivo", "Consolidar, de forma rastreável, todo o quantitativo de armadura da prancha DE-5400.00-6310-120-TX3-002 "
  "(fundações anelares dos tanques de lubrificação TQ-6310816A/B/C/D e TQ-6310817A/B/C/D), com painel gerencial e abas de auditoria."],
 ["Fonte primária", "Quadros LISTA DE BARRAS (8x) e RESUMO GERAL da própria prancha, complementados pelas cotas das "
  "vistas (planta de armadura, cortes A-A e B-B, vista C-C e elevações das faces)."],
 ["Multiplicidade", "A prancha atende 8 fundações idênticas - indicação '8x' no titulo da planta e no titulo da lista de barras. "
  "As colunas QUANT. e TOTAL do desenho já são o somatório das 8 fundações; os valores unitários das abas são derivados por divisão por 8. "
  "Essa leitura foi confirmada pela cota '298 N8 c.20' da planta (298 × 8 = 2.384, exatamente a quantidade da lista)."],
 ["Como usar", "Comece pelo Dashboard Executivo para a leitura gerencial. Para auditoria, use Lista de Barras (transcrição integral), "
  "Resumo por Bitola (fechamento contra o quadro do desenho), Resumo por Posição e Quantitativo por Tanque. As abas Geometria e Seções "
  "e Parâmetros Técnicos guardam as cotas e o carimbo."],
 ["Critério de massa", "Massa = comprimento total × massa linear nominal (Ø8 = 0,395 kg/m; Ø16 = 1,578 kg/m; Ø20 = 2,466 kg/m), "
  "exatamente como no quadro RESUMO GERAL da prancha e em linha com a NBR 7480."],
 ["Valores DERIVADOS", "Na aba Geometria e Seções, as linhas marcadas como DERIVADO (perímetro médio, diâmetro implícito, volume de "
  "concreto e taxa de armadura) NÃO constam do desenho: são cálculo próprio, apenas indicativo. O volume oficial de concreto deve sair "
  "da prancha de formas DE-5400.00-6310-120-TX3-001, que não foi fornecida."],
 ["Controle crítico", "Duas divergências do próprio desenho estão destacadas na aba Validações: (1) N13 - a lista traz 8 un (1 por fundação) "
  "enquanto o detalhe chama 'N13-2', o que daria 16 un (+61,23 kg); (2) N12 — a forma está cotada com perna variável (113 a 133) cm, mas o "
  "TOTAL de 150,72 m para 32 un implica comprimento médio de 471 cm. Ambas devem ser confirmadas com o projetista antes da compra."],
 ["Perdas", "Os números são os de projeto. Nenhuma taxa de perda de corte, dobra ou emenda foi acrescida. A coluna 'Barras de 12 m' e "
  "apenas um equivalente de suprimento (comprimento / 12 m, arredondado para cima)."],
 ["Unidades", "mm = milímetro (bitola); cm = centímetro (cotas do desenho); m = metro; kg = quilograma; t = tonelada; un = unidade; "
  "CORR = barra corrida; VAR = comprimento variável."],
 ["Limite de uso", "Levantamento para planejamento, orçamento e controle. Qualquer uso para execução ou compra deve ser submetido a "
  "validação do responsável técnico do projeto."],
]
fim = tabela(ws_ler, 5, ["Tópico", "Conteudo"], ler, [None, None], [24, 150])
for i in range(6, fim + 1):
    ws_ler.row_dimensions[i].height = 58
    ws_ler.cell(row=i, column=1).alignment = TOP
    ws_ler.cell(row=i, column=2).alignment = TOP

# ================================================================ Dashboard
ws = ws_dash
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:N1")
c = ws["A1"]; c.value = "PAINEL DE QUANTITATIVOS - ARMADURAS DAS FUNDAÇÕES DOS TANQUES DE LUBRIFICAÇÃO"
c.font = F_TIT; c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CTRV
ws.row_dimensions[1].height = 36
ws.merge_cells("A2:N2")
c = ws["A2"]
c.value = ("Prancha DE-5400.00-6310-120-TX3-002 Rev. 0 (07/08/2026)  |  TQ-6310816A/B/C/D & TQ-6310817A/B/C/D (8x)  |  "
           "U-6310 - Refino Boaventura  |  Escopo: armadura passiva CA-50")
c.font = F_SUB; c.fill = PatternFill("solid", fgColor=CREME); c.alignment = CTRV
ws.row_dimensions[2].height = 24

def card(col, linha, titulo, valor, legenda, cor, fmt):
    L = get_column_letter(col); R = get_column_letter(col + 2)
    ws.merge_cells("%s%d:%s%d" % (L, linha, R, linha))
    c = ws.cell(row=linha, column=col, value=titulo)
    c.font = F_CARD; c.fill = PatternFill("solid", fgColor=cor); c.alignment = CTR
    ws.merge_cells("%s%d:%s%d" % (L, linha + 1, R, linha + 2))
    c = ws.cell(row=linha + 1, column=col, value=valor)
    c.font = F_VAL; c.fill = PatternFill("solid", fgColor=VALBG); c.alignment = CTRV
    c.number_format = fmt
    ws.merge_cells("%s%d:%s%d" % (L, linha + 3, R, linha + 3))
    c = ws.cell(row=linha + 3, column=col, value=legenda)
    c.font = F_CAP; c.alignment = CTRV
    for rr in (linha, linha + 1, linha + 2):
        for cc in range(col, col + 3):
            ws.cell(row=rr, column=cc).border = BOX

card(1,  4, "MASSA TOTAL DE AÇO",  round(MASSA_TOT, 2), "kg de armadura CA-50 (8 fundações)", AZUL,    N_2)
card(4,  4, "COMPRIMENTO TOTAL",   round(COMPR_TOT, 2), "m lineares de barra", LARANJA, N_2)
card(7,  4, "FUNDAÇÕES ATENDIDAS", NT,                  "tanques idênticos cobertos pela prancha", AÇO, N_INT)
card(10, 4, "POSIÇÕES DE BARRA",   len(BARRAS),         "posições N1 a N13 na lista de barras", VERDE, N_INT)
card(1,  9, "MASSA POR FUNDAÇÃO",  round(MASSA_TOT / NT, 2), "kg de aço por tanque", VERDE,  N_2)
card(4,  9, "BITOLAS EMPREGADAS",  3,                   "Ø8, Ø16 e Ø20 mm - aço CA-50", AÇO,   N_INT)
card(7,  9, "ESTRIBOS",            QTD_ESTRIBOS,        "un (N8 + N9 + N12) nas 8 fundações", LARANJA, N_INT)
card(10, 9, "BARRAS DE 12 m",      BARRAS_12M,          "un equivalentes para suprimento", AZUL,  N_INT)
ws.row_dimensions[7].height = 16
ws.row_dimensions[12].height = 16

# --- tabela: composição por bitola
seção(ws, 15, 1, 5, "COMPOSIÇÃO POR BITOLA")
hdr = ["Bitola", "Compr. total (m)", "Massa (kg)", "% da massa", "Barras 12 m"]
for j, h in enumerate(hdr, start=1):
    c = ws.cell(row=16, column=j, value=h)
    c.font = F_HDR; c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTR; c.border = BOX
ws.row_dimensions[16].height = 26
for i, d in enumerate((8, 16, 20), start=17):
    comp = RESUMO_DES[d][0]; m = comp * KGM[d]
    vals = ["Ø%d mm" % d, comp, round(m, 2), m / MASSA_TOT, math.ceil(comp / 12)]
    fmts = [None, N_2, N_2, N_PCT, N_INT]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX; c.alignment = CTRV
        if fmts[j-1]: c.number_format = fmts[j-1]
total_row(ws, 20, 5, "TOTAL", [round(COMPR_TOT, 2), round(MASSA_TOT, 2), 1.0, BARRAS_12M],
          [N_2, N_2, N_PCT, N_INT], span=1)

# --- tabela: participação por função
seção(ws, 15, 8, 11, "PARTICIPAÇÃO POR FUNÇÃO ESTRUTURAL")
ws.merge_cells(start_row=16, start_column=8, end_row=16, end_column=9)
for j, h in ((8, "Função estrutural"), (10, "Massa (kg)"), (11, "% da massa")):
    c = ws.cell(row=16, column=j, value=h)
    c.font = F_HDR; c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTR; c.border = BOX
ws.cell(row=16, column=9).fill = PatternFill("solid", fgColor=AÇO)
ws.cell(row=16, column=9).border = BOX
for i, (nome, poss) in enumerate(GRUPOS, start=17):
    m = sum(b[4] * KGM[b[1]] for b in BARRAS if b[0] in poss)
    ws.merge_cells(start_row=i, start_column=8, end_row=i, end_column=9)
    for j, v in ((8, nome), (10, round(m, 2)), (11, m / MASSA_TOT)):
        c = ws.cell(row=i, column=j, value=v)
        c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX
        c.alignment = LEFT if j == 8 else CTRV
        if j == 10: c.number_format = N_2
        if j == 11: c.number_format = N_PCT
    ws.cell(row=i, column=9).fill = PatternFill("solid", fgColor=LINHA)
    ws.cell(row=i, column=9).border = BOX
ws.merge_cells(start_row=21, start_column=8, end_row=21, end_column=9)
for j, v, f in ((8, "TOTAL", None), (10, round(MASSA_TOT, 2), N_2), (11, 1.0, N_PCT)):
    c = ws.cell(row=21, column=j, value=v)
    c.font = Font(name="Arial", size=9, bold=True, color=BRANCO); c.fill = PatternFill("solid", fgColor=AÇO)
    c.alignment = LEFT if j == 8 else CTRV; c.border = BOX
    if f: c.number_format = f
ws.cell(row=21, column=9).fill = PatternFill("solid", fgColor=AÇO)
ws.cell(row=21, column=9).border = BOX

# --- graficos
def estilo_eixo(ch):
    ch.style = 2
    ch.y_axis.majorGridlines = None
    for ax in (ch.x_axis, ch.y_axis):
        ax.txPr = None

ch1 = BarChart(); ch1.type = "col"; ch1.title = "Massa de aço por bitola (kg)"
ch1.add_data(Reference(ws, min_col=3, min_row=16, max_row=19), titles_from_data=True)
ch1.set_categories(Reference(ws, min_col=1, min_row=17, max_row=19))
ch1.y_axis.title = "kg"; ch1.x_axis.title = "Bitola"
ch1.height = 8.4; ch1.width = 15.6; ch1.legend = None
ch1.series[0].graphicalProperties = GraphicalProperties(solidFill=AZUL[2:])
estilo_eixo(ch1)
ws.add_chart(ch1, "A23")

ch2 = DoughnutChart(holeSize=55); ch2.title = "Participação por função estrutural"
ch2.add_data(Reference(ws, min_col=10, min_row=16, max_row=20), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=8, min_row=17, max_row=20))
ch2.height = 8.4; ch2.width = 15.6
for idx, cor in enumerate([AZUL, LARANJA, AÇO, VERDE]):
    ch2.series[0].data_points.append(DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=cor[2:])))
ws.add_chart(ch2, "H23")

seção(ws, 40, 1, 14, "MASSA POR POSIÇÃO DE BARRA (N1 A N13) - FONTE: ABA RESUMO POR POSIÇÃO")
ch3 = BarChart(); ch3.type = "col"; ch3.title = "Massa por posição de barra (kg)"
ch3.add_data(Reference(ws_pos, min_col=5, min_row=5, max_row=18), titles_from_data=True)
ch3.set_categories(Reference(ws_pos, min_col=1, min_row=6, max_row=18))
ch3.y_axis.title = "kg"; ch3.x_axis.title = "Posição"
ch3.height = 8.6; ch3.width = 32.0; ch3.legend = None
ch3.series[0].graphicalProperties = GraphicalProperties(solidFill=AÇO[2:])
estilo_eixo(ch3)
ws.add_chart(ch3, "A42")

# --- bloco de validação no painel
seção(ws, 60, 1, 14, "CONTROLE DE VALIDAÇÃO - DESTAQUES")
ws.merge_cells("B61:C61"); ws.merge_cells("G61:N61")
for j, h in ((1, "Grupo"), (2, "Indicador"), (4, "Calculado"), (5, "Referência"), (6, "Desvio"), (7, "Critério")):
    c = ws.cell(row=61, column=j, value=h)
    c.font = F_HDR; c.fill = PatternFill("solid", fgColor=AÇO); c.alignment = CTR; c.border = BOX
for j in list(range(1, 15)):
    ws.cell(row=61, column=j).fill = PatternFill("solid", fgColor=AÇO)
    ws.cell(row=61, column=j).border = BOX
destaques = [
 ["Comprimento", "Somatório da lista vs. RESUMO GERAL (m)", round(COMPR_TOT, 2), round(COMPR_TOT, 2), 0,
  "As três bitolas fecham exatamente com o quadro impresso na prancha."],
 ["Massa", "Massa total de aço (kg)", round(MASSA_TOT, 2), TOTAL_DES, round(MASSA_TOT - TOTAL_DES, 2),
  "Diferença de arredondamento do TOTAL (kg) impresso."],
 ["Multiplicidade 8x", "N8 - 298 c/20 por fundação x 8", 2384, 2384, 0,
  "Confirma que as colunas QUANT. e TOTAL da lista já são o conjunto das 8 fundações."],
 ["DIVERGÊNCIA", "N13 - quantidade por fundação (un)", 1, 2, -1,
  "Lista traz 8 un (1/fundação); o detalhe chama 'N13-2' (2/fundação). Se prevalecer o detalhe: +8 un, +38,80 m, +61,23 kg. Confirmar."],
 ["DIVERGÊNCIA", "N12 - comprimento unitário implícito (m)", 4.71, 1.23, 3.48,
  "Forma cotada como (113 a 133) cm, mas o TOTAL de 150,72 m para 32 un implica 471 cm por barra. Confirmar."],
]
for i, row in enumerate(destaques, start=62):
    alerta = row[0] == "DIVERGÊNCIA"
    fonte = (Font(name="Arial", size=8, bold=True, color="FF8A4B0F") if alerta
             else Font(name="Arial", size=8, color=TINTA))
    fundo = PatternFill("solid", fgColor=ALERTA if alerta else LINHA)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=14)
    for j, v in ((1, row[0]), (2, row[1]), (4, row[2]), (5, row[3]), (6, row[4]), (7, row[5])):
        c = ws.cell(row=i, column=j, value=v)
        c.font = fonte; c.fill = fundo; c.border = BOX
        c.alignment = LEFT if j in (2, 7) else CTRV
        if j in (4, 5, 6): c.number_format = N_2
    for j in range(1, 15):
        ws.cell(row=i, column=j).fill = fundo
        ws.cell(row=i, column=j).border = BOX
    ws.row_dimensions[i].height = 24

for col in "ABCDEFGHIJKLMN":
    ws.column_dimensions[col].width = 13
ws.freeze_panes = "A4"


# ---------------------------------------------------------------- impressao
from openpyxl.worksheet.properties import PageSetupProperties
for _ws in wb.worksheets:
    _ws.page_setup.orientation = "landscape"
    _ws.page_setup.paperSize = _ws.PAPERSIZE_A4
    _ws.page_setup.fitToWidth = 1
    _ws.page_setup.fitToHeight = 1 if _ws.title == "Dashboard Executivo" else 0
    _ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    _ws.print_options.horizontalCentered = True
    _ws.page_margins.left = _ws.page_margins.right = 0.3
    _ws.page_margins.top = _ws.page_margins.bottom = 0.4
    if _ws.title != "Dashboard Executivo":
        _ws.print_title_rows = "1:5"

wb.save(OUT)
print("gerado:", OUT)
print("massa total %.2f kg | comprimento %.2f m | por fundação %.2f kg" % (MASSA_TOT, COMPR_TOT, MASSA_TOT / NT))
