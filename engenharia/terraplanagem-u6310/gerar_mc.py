# -*- coding: utf-8 -*-
"""Gera a Memoria de Calculo de Terraplenagem - Area dos Tanques U-6310 / U-6312."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/ProjectX-/engenharia/terraplanagem-u6310/MC-TERRAPLANAGEM_U-6310_U-6312_Rev0.xlsx"

AZ  = Font(name="Arial", size=10, color="0000FF")          # entrada (hardcoded)
PR  = Font(name="Arial", size=10)                          # formula
VD  = Font(name="Arial", size=10, color="008000")          # link outra aba
BOLD= Font(name="Arial", size=10, bold=True)
TIT = Font(name="Arial", size=14, bold=True, color="1F3864")
SUB = Font(name="Arial", size=11, bold=True, color="1F3864")
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SMALL=Font(name="Arial", size=8, italic=True, color="595959")
RED = Font(name="Arial", size=10, bold=True, color="C00000")

FH  = PatternFill("solid", fgColor="1F3864")
FA  = PatternFill("solid", fgColor="FFFF00")   # premissa chave / preencher
FT  = PatternFill("solid", fgColor="DDEBF7")   # totais
FW  = PatternFill("solid", fgColor="FCE4D6")   # atencao
thin= Side(style="thin", color="BFBFBF")
BOX = Border(left=thin,right=thin,top=thin,bottom=thin)

N2  = '#,##0.00'
N3  = '#,##0.000'
PCT = '0.00%'

wb = openpyxl.Workbook()

def head(ws, title, sub=None):
    ws["A1"] = title; ws["A1"].font = TIT
    if sub:
        ws["A2"] = sub; ws["A2"].font = SMALL
    ws.sheet_view.showGridLines = False

def hrow(ws, r, labels, widths=None, start=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=r, column=start+i, value=lab)
        c.font = HDR; c.fill = FH; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 30
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start+i)].width = w

def put(ws, r, c, v, font=PR, fmt=None, fill=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if border: cell.border = BOX
    if align: cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell

# =====================================================================
# 1) CAPA
# =====================================================================
ws = wb.active; ws.title = "CAPA"
head(ws, "MEMÓRIA DE CÁLCULO — TERRAPLENAGEM DE IMPLANTAÇÃO")
ws["A2"] = "Área dos Tanques — U-6310 (Parque de Tanques de Produtos Acabados / Lubrificantes) e U-6312 (UCO)"
ws["A2"].font = SUB
ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 74

dados = [
    ("IDENTIFICAÇÃO", None),
    ("Programa", "REFINO BOAVENTURA"),
    ("Cliente", "TR / BOAVENTURA — PETROBRAS"),
    ("Contrato", "5900.0131990.25.2 — SRGE/SI-III"),
    ("Executante", "CONSÓRCIO TEM BOAVENTURA"),
    ("Disciplina", "CIVIL — Terraplenagem"),
    ("Objeto", "Escavação, troca de solo, reaterro e compactação das bases dos tanques"),
    ("", None),
    ("DOCUMENTOS BASE DESTA MEMÓRIA", None),
    ("Planta (quantitativos)", "DE-5400.00-6310-113-TX3-003 rev.0 — U-6310 Planta de Terraplenagem de Implantação — Área dos Tanques (07/08/2026)"),
    ("Cortes / perfis", "DE-5400.00-6310-113-TX3-004 rev.0 — idem — Cortes A-A a E-E (07/08/2026)"),
    ("Especificação técnica", "ET-5400.00-6310-113-TX3-001 — Terraplenagem"),
    ("Relatório geotécnico", "RL-5400.00-6310-115-TX3-001 — Relatório Geotécnico Interpretativo (pendência P5)"),
    ("Planilha de fichas de tarefa", "IMPORTAÇÃO_TERRAPLAN_TANQUE.xlsx — aba 'IMPORTAÇÃO' (10 FTs, FT-CV-2619 a FT-CV-2628)"),
    ("Correio eletrônico", "Luciana Melo (Coord. Eng. Civil) — 11/08/2026 — 'EPC03-U-6310 / U-6312 — Terraplanagem de implantação dos tanques'"),
    ("", None),
    ("ESCOPO COBERTO", None),
    ("Tanques", "17 unidades: TQ-6310815A/B/C/D, 816A/B/C/D, 817A/B/C/D, 818A/B/C/D e TQ-6312824"),
    ("Atividades (conforme FT)", "001-CORTE; 002-TROCA DE SOLO; 003-ATERRO; 004-COMPACTAÇÃO"),
    ("", None),
    ("ORGANIZAÇÃO DO ARQUIVO", None),
    ("PREMISSAS", "Parâmetros de projeto. Todas as demais abas referenciam esta — alterou aqui, recalcula tudo."),
    ("GEOMETRIA", "Raios e cotas por linha de tanques, extraídos/aferidos nas plantas."),
    ("MC-TANQUE", "Memória de cálculo do volume de escavação, tanque a tanque (17 linhas)."),
    ("AFERIÇÃO", "Confronto do modelo geométrico contra a tabela de quantitativos do desenho."),
    ("QUANT-LINHA", "Consolidação por linha de tanques: corte, troca de solo e plataforma/bacia."),
    ("DISTRIB-FT", "Distribuição dos quantitativos pelas 10 fichas de tarefa (por tanque, não por área total)."),
    ("CONFRONTO-FT", "Planilha de importação vigente × esta memória de cálculo. Divergências apuradas."),
    ("RESUMO", "Quantitativos finais para produção, suprimentos e planejamento."),
]
r = 4
for k, v in dados:
    if v is None and k:
        ws.cell(row=r, column=1, value=k).font = SUB
    elif k or v:
        ws.cell(row=r, column=1, value=k).font = BOLD
        c = ws.cell(row=r, column=2, value=v); c.font = PR
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
ws.cell(row=r, column=1, value="CONVENÇÃO DE CORES").font = SUB; r += 1
for txt, fnt, fil in [("Azul — dado de entrada (cotado no desenho ou definido pelo usuário). Editável.", AZ, None),
                      ("Preto — resultado calculado por fórmula. Não editar.", PR, None),
                      ("Verde — valor trazido de outra aba.", VD, None),
                      ("Fundo amarelo — premissa chave a confirmar com o projetista.", BOLD, FA)]:
    c = ws.cell(row=r, column=1, value=txt); c.font = fnt
    if fil: c.fill = fil
    r += 1

r += 1
ws.cell(row=r, column=1, value="NOTA IMPORTANTE").font = RED
ws.cell(row=r, column=2, value=("Os volumes da tabela do desenho são GEOMÉTRICOS (sem empolamento) e foram calculados para a "
    "elevação de limpeza do terreno EL. 19,00. Conforme correio de 11/08/2026, o reaterro das bases será feito com "
    "substituição por rachão — logo o volume de 'REATERRO' do desenho é, na prática, BOTA-FORA de material escavado.")).font = RED
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 46

# =====================================================================
# 2) PREMISSAS
# =====================================================================
ws = wb.create_sheet("PREMISSAS")
head(ws, "PREMISSAS DE CÁLCULO", "Fonte: DE-5400.00-6310-113-TX3-003/004 rev.0 (notas gerais e cortes) e ET-5400.00-6310-113-TX3-001")
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 13
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 62

hrow(ws, 4, ["Item", "Parâmetro", "Valor", "Unid.", "Origem / justificativa"])
prem = [
 ("1",  "Cota do terreno regularizado (limpeza)", 19.00, "m", "Nota 7 dos cortes A-A a E-E e nota da tabela de quantitativos: 'volumes calculados considerando a elevação de limpeza do terreno (EL. 19,00)'.", True),
 ("2",  "Cota do greide de projeto — fundo sob tanque", 17.50, "m", "'GREIDE DE PROJETO EL.17,50' nos cortes A-A a E-E; 'EL.+17,50' no centro de cada tanque na planta.", True),
 ("3",  "Profundidade de escavação sob o tanque", None, "m", "Calculado: Item 1 − Item 2. Coerente com a Nota 4 (profundidade de 1,50 m da cota do terreno).", False),
 ("4",  "Talude de corte (H:V)", 1.00, "H:1V", "Nota 6: 'TALUDES DE CORTE - 1:1 (H:V)'. Confirmado pela marcação 1/1 nos cortes.", True),
 ("5",  "Projeção horizontal do talude de corte", None, "m", "Calculado: Item 3 × Item 4.", False),
 ("6",  "Sobrelargura de troca de solo além do costado", 4.00, "m", "Nota 4: 'remoção e substituição de solo em um raio de 4,00 m do costado do tanque'. Cota '4,00' repetida em todos os tanques na planta e nos cortes.", True),
 ("7",  "Talude de aterro (H:V)", 1.50, "H:1V", "Nota 6: 'TALUDES DE ATERRO - 1.5:1 (H:V)'. Não incide nos volumes desta memória (não há aterro de plataforma).", True),
 ("8",  "Cota ponto baixo da bacia — U-6310", 18.90, "m", "'EL.:18,90 PTO. BAIXO' em toda a área dos tanques 815/816/817/818.", True),
 ("9",  "Cota ponto alto da bacia — U-6310 / U-6312", 19.00, "m", "'EL.:19,00 PTO. ALTO'.", True),
 ("10", "Cota ponto baixo da bacia — U-6312 (UCO)", 18.70, "m", "'EL.:18,70 PTO. BAIXO' na bacia do TQ-6312824.", True),
 ("11", "Declividade mínima da bacia — U-6310", 0.015, "m/m", "Símbolo '≥1,5% (INC.)' na planta, área U-6310.", True),
 ("12", "Declividade mínima da bacia — U-6312", 0.010, "m/m", "Símbolo '≥1,0% (INC.)' na planta, área do TQ-6312824.", True),
 ("13", "Fator de empolamento (corte → solto)", 1.40, "—", "Fator vigente na planilha de importação de FTs (multiplicador 1,4). A CONFIRMAR com ensaio/ET — o estudo anterior (aba 'Planilha1') usava 1,30.", True),
 ("14", "Fator de empolamento — cenário alternativo", 1.30, "—", "Valor usado no estudo preliminar da equipe — aba 'Planilha1', célula B9 do arquivo original IMPORTAÇÃO_TERRAPLAN_TANQUE.xlsx. Mantido para análise de sensibilidade.", True),
 ("15", "Nº de fichas de tarefa (FT) por linha de tanques", 2, "un", "Critério da equipe: cada linha de 4 tanques é executada em 2 FTs (D/C e B/A). TQ-6312824 é dividido em 2 FTs (_1 e _2).", True),
]
r = 5
for it, nome, val, un, org, is_in in prem:
    put(ws, r, 1, it, BOLD, align="center")
    put(ws, r, 2, nome, PR)
    if val is None:
        f = "=C5-C6" if it == "3" else "=C7*C8"
        c = put(ws, r, 3, f, PR, N2, align="center")
    else:
        fmt = PCT if it in ("11", "12") else (N2 if isinstance(val, float) else '#,##0')
        c = put(ws, r, 3, val, AZ, fmt, FA if it in ("6", "13", "14") else None, align="center")
    put(ws, r, 4, un, PR, align="center")
    cc = put(ws, r, 5, org, SMALL)
    cc.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.cell(row=r, column=2, value="FORMULAÇÃO ADOTADA").font = SUB; r += 1
for t in [
  "A escavação sob cada tanque é um TRONCO DE CONE: base menor no fundo (EL. 17,50), base maior na superfície (EL. 19,00),",
  "gerado pelo talude de corte 1:1 ao longo dos 1,50 m de profundidade.",
  "",
  "        R_fundo  = R_costado + sobrelargura (4,00 m)          → é o raio cotado nas plantas (R13,71 / R10,85 / R18,53)",
  "        R_topo   = R_fundo + projeção do talude (1,50 m)      → é o raio cotado nas plantas (R15,21 / R12,35 / R20,03)",
  "        V        = (π · h / 3) · (R_fundo² + R_fundo · R_topo + R_topo²)",
  "",
  "Este modelo reproduz a coluna REATERRO da tabela de quantitativos do desenho com desvio de 0,25% no total",
  "(ver aba AFERIÇÃO), o que confirma a formulação e permite desagregar o quantitativo do desenho por tanque.",
]:
    c = ws.cell(row=r, column=2, value=t)
    c.font = Font(name="Consolas", size=9) if t.startswith("    ") else SMALL
    r += 1

# =====================================================================
# 3) GEOMETRIA
# =====================================================================
ws = wb.create_sheet("GEOMETRIA")
head(ws, "GEOMETRIA POR LINHA DE TANQUES", "Raios cotados em DE-5400.00-6310-113-TX3-003 rev.0. A linha 818 não tem cota de raio na planta — ver coluna 'Origem do raio'.")
hrow(ws, 4, ["Linha", "Tanques", "Produto", "Qtd\ntanques", "Espaç.\neixos (m)",
             "R fundo da\nescavação (m)", "R topo da\nescavação (m)", "R do costado\ndo tanque (m)",
             "Ø do tanque\n(m)", "R topo cotado\nno desenho (m)", "Origem do raio"],
     [9, 22, 14, 8, 9, 12, 12, 12, 10, 12, 46])

GEO = [
 ("816", "TQ-6310816A a D", "OB 100N",     4, 29.00, 13.71, 15.21, "Cotado na planta: R13,71 e R15,21"),
 ("817", "TQ-6310817A a D", "OB 220N",     4, 29.00, 13.71, 15.21, "Cotado na planta: R13,71 e R15,21"),
 ("818", "TQ-6310818A a D", "OB 500/600N", 4, 27.50, 12.78, None,  "Sem cota de raio na planta. Retro-calculado pelo volume do desenho (12,7818 m) e corroborado por medição independente no arquivo vetorial (12,77 m)."),
 ("815", "TQ-6310815A a D", "OB 80N",      4, 24.00, 10.85, 12.35, "Cotado na planta: R10,85 e R12,35"),
 ("824", "TQ-6312824",      "UCO",         1, None,  18.53, 20.03, "Cotado na planta: R18,53 e R20,03"),
]
r = 5
for lin, tq, prod, qt, esp, rf, rt_cot, org in GEO:
    put(ws, r, 1, lin, BOLD, align="center")
    put(ws, r, 2, tq, PR)
    put(ws, r, 3, prod, PR, align="center")
    put(ws, r, 4, qt, AZ, '#,##0', align="center")
    if esp: put(ws, r, 5, esp, AZ, N2, align="center")
    else:   put(ws, r, 5, "—", PR, align="center")
    put(ws, r, 6, rf, AZ, N2, align="center")
    put(ws, r, 7, f"=F{r}+PREMISSAS!$C$9", PR, N2, align="center")
    put(ws, r, 8, f"=F{r}-PREMISSAS!$C$10", PR, N2, align="center")
    put(ws, r, 9, f"=2*H{r}", PR, N2, align="center")
    if rt_cot: put(ws, r,10, rt_cot, AZ, N2, align="center")
    else:      put(ws, r,10, "n/d", PR, align="center")
    c = put(ws, r, 11, org, SMALL); c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

r += 1
ws.cell(row=r, column=1, value="Verificação da Nota 4 (sobrelargura de 4,00 m):").font = BOLD
r += 1
for t in ["O círculo tracejado 'COSTADO DO TANQUE' do TQ-6312824 foi medido no arquivo vetorial em R = 14,53 m.",
          "R fundo cotado (18,53 m) − 14,53 m = 4,00 m, exatamente a sobrelargura da Nota 4. A relação R_fundo = R_costado + 4,00 fica comprovada.",
          "A largura da faixa hachurada em azul (talude) mede 1,50 m nas plantas, igual à projeção horizontal do talude 1:1 sobre 1,50 m de altura."]:
    ws.cell(row=r, column=1, value=t).font = SMALL; r += 1

# =====================================================================
# 4) MC-TANQUE
# =====================================================================
ws = wb.create_sheet("MC-TANQUE")
head(ws, "MEMÓRIA DE CÁLCULO — ESCAVAÇÃO / TROCA DE SOLO POR TANQUE",
     "V = (π · h / 3) · (R_fundo² + R_fundo · R_topo + R_topo²)   —   tronco de cone entre EL. 17,50 e EL. 19,00, talude de corte 1:1")
hrow(ws, 4, ["#", "TAG do tanque", "Linha", "Unid.", "Produto", "R fundo\n(m)", "R topo\n(m)",
             "h\n(m)", "Área do fundo\n(m²)", "Área do topo\n(m²)",
             "Volume de escavação\n/ troca de solo (m³)", "Volume solto\n(× empolamento) (m³)"],
     [5, 18, 8, 8, 13, 9, 9, 7, 12, 12, 19, 16])

TANQUES = []
for lin, base, unid, prod, grow in [
    ("816", "TQ-6310816", "U-6310", "OB 100N", 5),
    ("817", "TQ-6310817", "U-6310", "OB 220N", 6),
    ("818", "TQ-6310818", "U-6310", "OB 500/600N", 7),
    ("815", "TQ-6310815", "U-6310", "OB 80N", 8),
]:
    for suf in ("A", "B", "C", "D"):
        TANQUES.append((base + suf, lin, unid, prod, grow))
TANQUES.append(("TQ-6312824", "824", "U-6312", "UCO", 9))

r = 5
for i, (tag, lin, unid, prod, grow) in enumerate(TANQUES, start=1):
    put(ws, r, 1, i, PR, '#,##0', align="center")
    put(ws, r, 2, tag, BOLD)
    put(ws, r, 3, lin, PR, align="center")
    put(ws, r, 4, unid, PR, align="center")
    put(ws, r, 5, prod, PR, align="center")
    put(ws, r, 6, f"=GEOMETRIA!$F${grow}", VD, N2, align="center")
    put(ws, r, 7, f"=GEOMETRIA!$G${grow}", VD, N2, align="center")
    put(ws, r, 8, "=PREMISSAS!$C$7", VD, N2, align="center")
    put(ws, r, 9, f"=PI()*F{r}^2", PR, N2)
    put(ws, r,10, f"=PI()*G{r}^2", PR, N2)
    put(ws, r,11, f"=PI()*H{r}/3*(F{r}^2+F{r}*G{r}+G{r}^2)", PR, N2)
    put(ws, r,12, f"=K{r}*PREMISSAS!$C$17", PR, N2)
    r += 1
first, last = 5, r - 1
put(ws, r, 2, "TOTAL — 17 tanques", BOLD, fill=FT)
for c in (1, 3, 4, 5, 6, 7, 8):
    put(ws, r, c, None, BOLD, fill=FT)
put(ws, r, 9,  f"=SUM(I{first}:I{last})", BOLD, N2, FT)
put(ws, r,10,  f"=SUM(J{first}:J{last})", BOLD, N2, FT)
put(ws, r,11,  f"=SUM(K{first}:K{last})", BOLD, N2, FT)
put(ws, r,12,  f"=SUM(L{first}:L{last})", BOLD, N2, FT)
MCT_TOTAL = r
r += 2
ws.cell(row=r, column=2, value=("Este volume é integralmente TROCA DE SOLO (Nota 4): o material escavado sai e é substituído por rachão, "
    "com posterior reaterro e controle de compactação. A escavação deve ser acompanhada por engenheiro geotécnico "
    "para ajuste da profundidade às condições reais do terreno.")).font = SMALL
r += 1
ws.cell(row=r, column=2, value="Não inclui o corte de regularização da plataforma/bacia — ver aba QUANT-LINHA.").font = SMALL

# =====================================================================
# 5) AFERICAO
# =====================================================================
ws = wb.create_sheet("AFERIÇÃO")
head(ws, "AFERIÇÃO DO MODELO CONTRA A TABELA DE QUANTITATIVOS DO DESENHO",
     "Tabela de quantitativos de escavação das áreas dos tanques — DE-5400.00-6310-113-TX3-003 rev.0")
hrow(ws, 4, ["Linha", "Local (conforme desenho)", "CORTE do\ndesenho (m³)", "REATERRO do\ndesenho (m³)",
             "Escavação dos tanques\ncalculada (m³)", "Desvio\n(m³)", "Desvio\n(%)", "Aderência"],
     [9, 40, 14, 14, 20, 12, 11, 14])

AFER = [("816", "TQ-6310816A ATÉ TQ-6310816D E BACIAS", 4100.16, 3945.84, 5,  8),
        ("817", "TQ-6310817A ATÉ TQ-6310817D E BACIAS", 4130.06, 3947.50, 9, 12),
        ("818", "TQ-6310818A ATÉ TQ-6310818D E BACIAS", 3599.36, 3455.05, 13,16),
        ("815", "TQ-6310815A ATÉ TQ-6310815D E BACIAS", 2649.19, 2492.22, 17,20),
        ("824", "TQ-6312824 E BACIA",                   2789.14, 1756.44, 21,21)]
r = 5
for lin, local, corte, reat, a, b in AFER:
    put(ws, r, 1, lin, BOLD, align="center")
    put(ws, r, 2, local, PR)
    put(ws, r, 3, corte, AZ, N2)
    put(ws, r, 4, reat,  AZ, N2)
    put(ws, r, 5, f"=SUM('MC-TANQUE'!K{a}:K{b})", VD, N2)
    put(ws, r, 6, f"=E{r}-D{r}", PR, N2)
    put(ws, r, 7, f"=IF(D{r}=0,\"\",F{r}/D{r})", PR, PCT)
    put(ws, r, 8, f'=IF(ABS(G{r})<=0.01,"Excelente",IF(ABS(G{r})<=0.03,"Boa","Verificar"))', PR, align="center")
    r += 1
put(ws, r, 2, "TOTAL", BOLD, fill=FT)
put(ws, r, 1, None, BOLD, fill=FT)
put(ws, r, 3, f"=SUM(C5:C{r-1})", BOLD, N2, FT)
put(ws, r, 4, f"=SUM(D5:D{r-1})", BOLD, N2, FT)
put(ws, r, 5, f"=SUM(E5:E{r-1})", BOLD, N2, FT)
put(ws, r, 6, f"=E{r}-D{r}", BOLD, N2, FT)
put(ws, r, 7, f"=F{r}/D{r}", BOLD, PCT, FT)
put(ws, r, 8, None, BOLD, fill=FT)
AF_TOT = r
r += 2
ws.cell(row=r, column=2, value="LEITURA DO RESULTADO").font = SUB; r += 1
for t in [
 "1. O modelo geométrico reproduz a coluna REATERRO do desenho com desvio total de +0,25%, e abaixo de 0,1% em três das cinco linhas (816, 817 e 818).",
 "   Isso comprova que o REATERRO do desenho corresponde ao volume escavado sob os tanques (tronco de cone, EL. 17,50 a 19,00, talude de corte 1:1).",
 "   Ressalva de independência: o raio da linha 818 foi retro-calculado a partir do próprio volume do desenho, de modo que o desvio de 0,03% dessa linha",
 "   não constitui verificação independente. A aferição é independente nas linhas 816, 817 e 824, cujos raios estão cotados na planta.",
 "",
 "2. A linha 815 é o único desvio relevante (+47,71 m³, +1,91%). Mecanismos identificados e quantificados por integração:",
 "        truncamento dos quatro cones pelo limite sul da plataforma (cota 10,97 m do eixo, menor que o raio de topo 12,35 m) .... 23,05 m³",
 "        truncamento do cone da extremidade oeste (cota 11,85 m) ........................................................  0,46 m³",
 "        truncamento do cone da extremidade leste (cota 11,15 m) ........................................................  4,08 m³",
 "        sobreposição entre cones adjacentes (espaçamento 24,00 m, raio de topo 12,35 m) ................................  1,14 m³",
 "        soma .......................................................................................................... 28,73 m³  (60% do desvio)",
 "   O saldo de aproximadamente 19 m³ (0,76% da linha) permanece em aberto e deve ser confirmado com o projetista.",
 "   Adota-se o valor do desenho por prevalência contratual do documento emitido — e não por margem de segurança: 2.492,22 m³ é o MENOR dos dois valores.",
 "",
 "3. Inconsistência interna do projeto a esclarecer: o corte D-D cota os espaçamentos de eixo da linha 815 como 24,11 / 23,89 / 24,00 m, enquanto a planta",
 "   e as coordenadas PDMS do próprio corte (E=4208,387 / 4232,387 / 4256,387 / 4280,387) dão 24,000 m exatos. Não afeta o volume, porque a soma de",
 "   troncos de cone independentes não depende do espaçamento entre eles.",
 "",
 "4. A diferença CORTE − REATERRO é o corte de regularização da plataforma e da bacia de contenção — quantificado na aba QUANT-LINHA.",
]:
    ws.cell(row=r, column=2, value=t).font = SMALL; r += 1

# =====================================================================
# 6) QUANT-LINHA
# =====================================================================
ws = wb.create_sheet("QUANT-LINHA")
head(ws, "CONSOLIDAÇÃO POR LINHA DE TANQUES", "Decomposição do CORTE do desenho em: escavação/troca de solo sob os tanques + corte de regularização da plataforma e bacia")
hrow(ws, 4, ["Linha", "Local", "Qtd\ntanques", "CORTE total\n(m³)", "Troca de solo\n— tanques (m³)",
             "Plataforma e bacia\n— corte compl. (m³)", "% tanques", "% bacia",
             "CORTE por\ntanque (m³)", "Troca de solo\npor tanque (m³)", "Bacia por\ntanque (m³)"],
     [9, 34, 8, 13, 15, 17, 10, 10, 13, 14, 12])
r = 5
for i, (lin, local, corte, reat, a, b) in enumerate(AFER):
    gr = 5 + i
    af = 5 + i
    put(ws, r, 1, lin, BOLD, align="center")
    put(ws, r, 2, local, PR)
    put(ws, r, 3, f"=GEOMETRIA!$D${gr}", VD, '#,##0', align="center")
    put(ws, r, 4, f"=AFERIÇÃO!C{af}", VD, N2)
    put(ws, r, 5, f"=AFERIÇÃO!D{af}", VD, N2)
    put(ws, r, 6, f"=D{r}-E{r}", PR, N2)
    put(ws, r, 7, f"=E{r}/D{r}", PR, PCT)
    put(ws, r, 8, f"=F{r}/D{r}", PR, PCT)
    put(ws, r, 9, f"=D{r}/C{r}", PR, N2)
    put(ws, r,10, f"=E{r}/C{r}", PR, N2)
    put(ws, r,11, f"=F{r}/C{r}", PR, N2)
    r += 1
put(ws, r, 2, "TOTAL", BOLD, fill=FT); put(ws, r, 1, None, BOLD, fill=FT)
put(ws, r, 3, f"=SUM(C5:C{r-1})", BOLD, '#,##0', FT)
for col in (4, 5, 6):
    put(ws, r, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})", BOLD, N2, FT)
put(ws, r, 7, f"=E{r}/D{r}", BOLD, PCT, FT)
put(ws, r, 8, f"=F{r}/D{r}", BOLD, PCT, FT)
for col in (9, 10, 11):
    put(ws, r, col, None, BOLD, fill=FT)
QL_TOT = r
r += 2
ws.cell(row=r, column=2, value="OBSERVAÇÃO SOBRE O RATEIO TANQUE / BACIA").font = SUB; r += 1
for t in [
 "A planilha de importação de FTs vigente rateia o volume em 70% para a 'área tanque' e 30% para a 'área bacia'.",
 "Pela geometria do projeto esse rateio não se sustenta nas linhas de U-6310: a escavação sob os tanques responde por 94% a 96% do corte,",
 "porque a bacia é apenas regularizada (EL. 19,00 → 18,90, i ≥ 1,5%) enquanto o tanque é escavado 1,50 m.",
 "Somente o TQ-6312824 se aproxima do rateio adotado (63% / 37%), por ter bacia bem maior e ponto baixo em EL. 18,70.",
 "Consequência: aplicar o fator 0,7 sobre o volume da linha subestima o escopo de escavação em cerca de 26 pontos percentuais.",
]:
    ws.cell(row=r, column=2, value=t).font = SMALL; r += 1

# =====================================================================
# 7) DISTRIB-FT
# =====================================================================
ws = wb.create_sheet("DISTRIB-FT")
head(ws, "DISTRIBUIÇÃO DOS QUANTITATIVOS PELAS FICHAS DE TAREFA",
     "Rateio por tanque (e não pela área total), preservando o particionamento adotado pela equipe: 2 FTs por linha de tanques")
hrow(ws, 4, ["Chave de\nimportação", "TAG do componente", "Data\nprogramada", "Unid.", "Código Primavera",
             "Linha", "Tanques abrangidos", "Nº de\ntanques", "Fração da linha\n(1 ÷ nº de FTs)",
             "CORTE geométrico\n(m³)", "Troca de solo\ngeométrico (m³)", "Plataforma/bacia\ngeométrico (m³)",
             "CORTE solto —\ntransporte (m³)", "Rachão a fornecer\n(m³ in situ)"],
     [12, 19, 12, 8, 20, 8, 21, 8, 8, 15, 14, 15, 14, 15])

import datetime as _dt
FTS = [
 ("FT-CV-2619", "TQ-6310816D/C", _dt.date(2026,8,28),  "U6310", "U6310-C-610246000", "816", "TQ-6310816D e TQ-6310816C", 2, 4),
 ("FT-CV-2620", "TQ-6310816B/A", _dt.date(2026,9,4),   "U6310", "U6310-C-610246000", "816", "TQ-6310816B e TQ-6310816A", 2, 4),
 ("FT-CV-2621", "TQ-6310817D/C", _dt.date(2026,9,11),  "U6310", "U6310-C-610250000", "817", "TQ-6310817D e TQ-6310817C", 2, 4),
 ("FT-CV-2622", "TQ-6310817B/A", _dt.date(2026,9,18),  "U6310", "U6310-C-610250000", "817", "TQ-6310817B e TQ-6310817A", 2, 4),
 ("FT-CV-2623", "TQ-6312824_1",  _dt.date(2026,9,25),  "U6312", "U6312-C-710073000", "824", "TQ-6312824 — 1ª metade",    1, 2),
 ("FT-CV-2624", "TQ-6312824_2",  _dt.date(2026,10,2),  "U6312", "U6312-C-710073000", "824", "TQ-6312824 — 2ª metade",    1, 2),
 ("FT-CV-2625", "TQ-6310818D/C", _dt.date(2026,10,9),  "U6310", "U6310-C-620256000", "818", "TQ-6310818D e TQ-6310818C", 2, 4),
 ("FT-CV-2626", "TQ-6310818B/A", _dt.date(2026,10,16), "U6310", "U6310-C-620256000", "818", "TQ-6310818B e TQ-6310818A", 2, 4),
 ("FT-CV-2627", "TQ-6310815D/C", _dt.date(2026,10,23), "U6310", "U6310-C-620252000", "815", "TQ-6310815D e TQ-6310815C", 2, 4),
 ("FT-CV-2628", "TQ-6310815B/A", _dt.date(2026,10,30), "U6310", "U6310-C-620252000", "815", "TQ-6310815B e TQ-6310815A", 2, 4),
]
QLD = "'QUANT-LINHA'!$D$5:$D$9"; QLE = "'QUANT-LINHA'!$E$5:$E$9"
QLF = "'QUANT-LINHA'!$F$5:$F$9"; QLA = "'QUANT-LINHA'!$A$5:$A$9"
r = 5
for chave, tag, data, unid, prim, lin, abr, nt, ntot in FTS:
    put(ws, r, 1, chave, BOLD, align="center")
    put(ws, r, 2, tag, PR)
    c = put(ws, r, 3, data, AZ, 'dd/mm/yyyy', align="center")
    put(ws, r, 4, unid, PR, align="center")
    put(ws, r, 5, prim, PR, align="center")
    put(ws, r, 6, lin, PR, align="center")
    put(ws, r, 7, abr, PR)
    put(ws, r, 8, nt, AZ, '#,##0', align="center")
    put(ws, r, 9, "=1/PREMISSAS!$C$19", PR, '0.0%', align="center")
    put(ws, r,10, f"=INDEX({QLD},MATCH($F{r},{QLA},0))*$I{r}", PR, N2)
    put(ws, r,11, f"=INDEX({QLE},MATCH($F{r},{QLA},0))*$I{r}", PR, N2)
    put(ws, r,12, f"=INDEX({QLF},MATCH($F{r},{QLA},0))*$I{r}", PR, N2)
    put(ws, r,13, f"=J{r}*PREMISSAS!$C$17", PR, N2)
    put(ws, r,14, f"=K{r}", PR, N2)
    r += 1
put(ws, r, 1, "TOTAL", BOLD, fill=FT)
for c in range(2, 10): put(ws, r, c, None, BOLD, fill=FT)
for col in (10, 11, 12, 13, 14):
    L = get_column_letter(col)
    put(ws, r, col, f"=SUM({L}5:{L}{r-1})", BOLD, N2, FT)
DF_TOT = r
r += 2
for t in [
 "Coluna 'CORTE geométrico' = corte total da linha (tabela do desenho) × fração de tanques da FT. É o volume no corte, sem empolamento.",
 "Coluna 'CORTE solto' = volume geométrico × fator de empolamento da aba PREMISSAS (item 13). É o volume a transportar / medir em caçamba.",
 "Coluna 'Rachão a fornecer' = volume da troca de solo, in situ compactado. O fator de conversão para volume solto de fornecimento deve ser",
 "definido com o fornecedor / ET antes da compra; não está aplicado aqui.",
 "As datas programadas reproduzem a planilha de importação vigente e não foram alteradas por esta memória de cálculo.",
 "Cada linha de tanques é executada em 2 FTs, logo a fração é sempre 1/2 (PREMISSAS, item 15). Para a linha 816, por exemplo, cada FT abrange 2 dos 4 tanques;",
 "para o TQ-6312824, tanque único, cada FT abrange metade do serviço do mesmo tanque.",
]:
    ws.cell(row=r, column=1, value=t).font = SMALL; r += 1

# =====================================================================
# 8) CONFRONTO-FT
# =====================================================================
ws = wb.create_sheet("CONFRONTO-FT")
head(ws, "CONFRONTO — PLANILHA DE IMPORTAÇÃO VIGENTE × MEMÓRIA DE CÁLCULO",
     "Planilha IMPORTAÇÃO_TERRAPLAN_TANQUE.xlsx, aba 'IMPORTAÇÃO', coluna I (Quantidade)")
for col, w in zip("ABCDEFGHIJ", [12, 19, 8, 26, 13, 15, 16, 13, 11, 40]):
    ws.column_dimensions[col].width = w

ws["A4"] = "Fatores empregados na planilha vigente"; ws["A4"].font = SUB
ws["A5"] = "Divisor de partição (metade da linha)"; ws["A5"].font = PR
ws["C5"] = 2; ws["C5"].font = AZ; ws["C5"].alignment = Alignment(horizontal="center")
ws["A6"] = "Fator 'por área tanque'"; ws["A6"].font = PR
ws["C6"] = 0.70; ws["C6"].font = AZ; ws["C6"].number_format = PCT; ws["C6"].fill = FW
ws["A7"] = "Fator de empolamento"; ws["A7"].font = PR
ws["C7"] = 1.40; ws["C7"].font = AZ; ws["C7"].number_format = N2
ws["D5"] = "Fórmula original da planilha:  =(BASE/2)*0,7*1,4"; ws["D5"].font = SMALL
ws["D6"] = "O fator 0,7 não corresponde à geometria do projeto — ver aba QUANT-LINHA."; ws["D6"].font = SMALL

hrow(ws, 9, ["Chave de\nimportação", "TAG do componente", "Linha", "Fórmula vigente\n(planilha)",
             "Base adotada\nna planilha", "Quantidade\nvigente (m³)", "Quantidade desta\nMC — solto (m³)",
             "Diferença\n(vigente − MC)", "Desvio\n(%)", "Origem da base adotada na planilha"])
BASES = {"816": 2434.25, "817": 2464.06, "818": 2071.27, "815": 1415.86, "824": 2235.88}
r = 10
for i, (chave, tag, data, unid, prim, lin, abr, nt, ntot) in enumerate(FTS):
    dfr = 5 + i
    put(ws, r, 1, chave, BOLD, align="center")
    put(ws, r, 2, tag, PR)
    put(ws, r, 3, lin, PR, align="center")
    put(ws, r, 4, f'=(TEXT(E{r},"#,##0.00")&" ÷ "&TEXT($C$5,"0")&" × "&TEXT($C$6,"0%")&" × "&TEXT($C$7,"0.00"))', PR, align="center")
    put(ws, r, 5, BASES[lin], AZ, N2)
    put(ws, r, 6, f"=E{r}/$C$5*$C$6*$C$7", PR, N2)
    put(ws, r, 7, f"='DISTRIB-FT'!M{dfr}", VD, N2)
    put(ws, r, 8, f"=F{r}-G{r}", PR, N2)
    put(ws, r, 9, f"=F{r}/G{r}-1", PR, PCT)
    put(ws, r,10, "Não rastreável aos desenhos rev.0 — ver nota abaixo", SMALL)
    r += 1
put(ws, r, 1, "TOTAL", BOLD, fill=FT)
for c in (2, 3, 4, 5, 10): put(ws, r, c, None, BOLD, fill=FT)
for col in (6, 7, 8):
    L = get_column_letter(col)
    put(ws, r, col, f"=SUM({L}10:{L}{r-1})", BOLD, N2, FT)
put(ws, r, 9, f"=F{r}/G{r}-1", BOLD, PCT, FT)
CF_TOT = r
r += 2
ws.cell(row=r, column=1, value="APURAÇÃO DAS DIVERGÊNCIAS").font = SUB; r += 1
for t in [
 "1) BASE NÃO RASTREÁVEL. As bases usadas na planilha (2.434,25 / 2.464,06 / 2.071,27 / 1.415,86 / 2.235,88 m³) não constam de nenhuma das",
 "   duas plantas emitidas em 07/08/2026 e não se obtêm por operação sobre as colunas CORTE ou REATERRO da tabela de quantitativos.",
 "   Representam de 53,4% a 80,2% do corte da linha correspondente, sem proporção constante. Provável origem: quantitativo preliminar,",
 "   anterior à emissão rev.0 — coerente com a observação registrada na própria planilha ('o ID usado sofrerá alterações com a aprovação do Book B').",
 "",
 "2) DUPLA REDUÇÃO. Sobre essa base já reduzida ainda se aplica o fator 0,7 ('por área tanque'), que pela geometria do projeto deveria ser",
 "   de 94% a 96% nas linhas de U-6310 (aba QUANT-LINHA). As duas reduções se somam.",
 "",
 "3) EFEITO. O total das 10 fichas fica em 10.408,89 m³ contra 24.175,07 m³ de corte empolado apurado nesta memória — 43,1% do escopo de corte.",
 "   O mesmo percentual vale em volume geométrico (10.408,89 ÷ 1,40 = 7.434,92 m³ contra 17.267,91 m³), já que o fator de empolamento incide nos dois lados.",
 "",
 "4) ENCAMINHAMENTO. Substituir a coluna 'Quantidade' da planilha de importação pelos valores da coluna G desta aba antes de importar",
 "   para o Primavera, e confirmar com a Coordenação de Engenharia Civil o fator de empolamento a adotar (1,30 ou 1,40).",
 "",
 "5) ALCANCE DA COLUNA 'Quantidade'. Tanto na planilha vigente quanto nesta memória a quantidade da FT mede APENAS o corte (atividade 001).",
 "   As atividades 002-TROCA DE SOLO, 003-ATERRO e 004-COMPACTAÇÃO, também declaradas na ficha, correspondem a outros 15.597,05 m³ in situ",
 "   que não estão medidos em nenhuma coluna da planilha de importação. Convenção mantida igual à do cliente — registrada aqui para não induzir a erro.",
]:
    c = ws.cell(row=r, column=1, value=t)
    c.font = RED if t.startswith(("1)", "2)", "3)", "4)")) else SMALL
    r += 1

# =====================================================================
# 9) RESUMO
# =====================================================================
ws = wb.create_sheet("RESUMO")
head(ws, "RESUMO DOS QUANTITATIVOS DE TERRAPLENAGEM", "Área dos tanques U-6310 e U-6312 — base: DE-5400.00-6310-113-TX3-003/004 rev.0")
for col, w in zip("ABCDE", [6, 56, 16, 10, 62]):
    ws.column_dimensions[col].width = w

hrow(ws, 4, ["Item", "Serviço", "Quantidade", "Unid.", "Critério / observação"])
RES = [
 ("1", "Escavação — corte total (volume geométrico)", f"='QUANT-LINHA'!D{QL_TOT}", "m³",
  "Soma da coluna CORTE da tabela de quantitativos do desenho. Sem empolamento, referida à EL. de limpeza 19,00."),
 ("1.1", "     Escavação / troca de solo sob os tanques", f"='QUANT-LINHA'!E{QL_TOT}", "m³",
  "Valor da coluna REATERRO da tabela do desenho (adotado). O modelo de tronco de cone desta memória o reproduz com +0,25% — ver AFERIÇÃO."),
 ("1.2", "     Corte de regularização da plataforma e bacia", f"='QUANT-LINHA'!F{QL_TOT}", "m³",
  "Diferença CORTE − REATERRO do desenho. Inclui a conformação da bacia (EL. 19,00 → 18,90 em U-6310 e → 18,70 em U-6312)."),
 ("2", "Escavação — volume solto para transporte", f"=C5*PREMISSAS!$C$17", "m³",
  "Item 1 × fator de empolamento (PREMISSAS, item 13). Volume de medição em caçamba."),
 ("2.1", "     Cenário alternativo (fator 1,30)", f"=C5*PREMISSAS!$C$18", "m³",
  "Sensibilidade — fator do estudo preliminar da equipe."),
 ("3", "Bota-fora", f"=C5*PREMISSAS!$C$17", "m³",
  "Correio de 11/08/2026: o reaterro das bases será feito com substituição por rachão, logo todo o material escavado é bota-fora. Volume solto."),
 ("4", "Rachão para reaterro das bases (volume in situ)", f"='QUANT-LINHA'!E{QL_TOT}", "m³",
  "Igual ao item 1.1. Fator de conversão para volume de fornecimento a definir com o fornecedor / ET."),
 ("5", "Reaterro e compactação com controle tecnológico", f"='QUANT-LINHA'!E{QL_TOT}", "m³",
  "Executar conforme RL-5400.00-6310-115-TX3-001 e ET-5400.00-6310-113-TX3-001. Controle tecnológico obrigatório (Nota 7)."),
 ("6", "Nº de tanques", "=SUM(GEOMETRIA!D5:D9)", "un",
  "16 tanques em U-6310 (4 linhas × 4) + 1 tanque UCO em U-6312."),
 ("7", "Nº de fichas de tarefa", "=COUNTA('DISTRIB-FT'!A5:A14)", "un",
  "Conforme planilha de importação vigente: FT-CV-2619 a FT-CV-2628."),
]
r = 5
for it, serv, qt, un, obs in RES:
    put(ws, r, 1, it, BOLD, align="center")
    c = put(ws, r, 2, serv, BOLD if "." not in it else PR)
    put(ws, r, 3, qt, VD if isinstance(qt, str) else AZ, N2 if un == "m³" else '#,##0')
    put(ws, r, 4, un, PR, align="center")
    o = put(ws, r, 5, obs, SMALL); o.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.cell(row=r, column=2, value="CONFRONTO COM A PLANILHA DE IMPORTAÇÃO VIGENTE").font = SUB; r += 1
hrow(ws, r, ["", "Indicador", "Valor", "Unid.", "Observação"]); r += 1
CONF = [
 ("Quantidade total lançada nas 10 FTs (planilha vigente)", f"='CONFRONTO-FT'!F{CF_TOT}", "m³", "Soma da coluna 'Quantidade' da aba IMPORTAÇÃO."),
 ("Quantidade total apurada nesta memória de cálculo", f"='CONFRONTO-FT'!G{CF_TOT}", "m³", "Corte empolado, distribuído por tanque."),
 ("Escopo faltante nas fichas de tarefa", f"='CONFRONTO-FT'!G{CF_TOT}-'CONFRONTO-FT'!F{CF_TOT}", "m³", "Diferença a lançar antes da importação para o Primavera."),
 ("Cobertura da planilha vigente", f"='CONFRONTO-FT'!F{CF_TOT}/'CONFRONTO-FT'!G{CF_TOT}", "—", "Percentual do escopo efetivamente lançado."),
]
for nome, f, un, obs in CONF:
    put(ws, r, 1, None, PR)
    put(ws, r, 2, nome, BOLD)
    put(ws, r, 3, f, PR, PCT if un == "—" else N2, FW)
    put(ws, r, 4, un, PR, align="center")
    o = put(ws, r, 5, obs, SMALL); o.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

r += 1
ws.cell(row=r, column=2, value="RESSALVAS E PENDÊNCIAS").font = SUB; r += 1
for t in [
 "a) Pendência P5 do desenho — RL-5400.00-6310-115-TX3-001 (Relatório Geotécnico Interpretativo). A profundidade de substituição de solo",
 "   deverá ser ajustada às condições reais do terreno, com acompanhamento de engenheiro geotécnico (Nota 4). Volumes desta memória",
 "   valem para a profundidade de projeto de 1,50 m.",
 "b) Fator de empolamento (1,40) não tem respaldo documental nos desenhos — deve ser confirmado por ensaio ou pela ET antes da medição.",
 "c) Raio da linha 818 não está cotado na planta. Adotado 12,78 m, retro-calculado a partir do volume do desenho e corroborado por medição no arquivo",
 "   vetorial (12,77 m). Confirmar a cota com o projetista — a aferição dessa linha não é independente.",
 "c1) Linha 815: desvio de +47,71 m³ (+1,91%) explicado em 60% por truncamento e sobreposição dos cones (ver AFERIÇÃO); saldo a confirmar com o projetista.",
 "c2) Corte D-D cota os espaçamentos da linha 815 como 24,11 / 23,89 / 24,00 m, contra 24,000 m exatos na planta e nas coordenadas PDMS. A esclarecer.",
 "d) A tabela do desenho não distingue o volume do dique e dos mini-diques de contenção mostrados nos cortes A-A a E-E; se forem executados",
 "   em aterro compactado, o material e o volume correspondentes devem ser orçados à parte.",
 "e) Não há coluna de aterro na tabela do desenho. Caso parte do corte venha a ser reaproveitada em outra frente, o bota-fora do item 3 reduz",
 "   na mesma proporção.",
 "f) A coluna 'Quantidade' das fichas de tarefa mede apenas o corte (atividade 001). Troca de solo, aterro e compactação (itens 4 e 5 acima) não estão",
 "   medidos em nenhuma coluna da planilha de importação — convenção herdada do cliente, registrada aqui para não induzir a erro.",
]:
    ws.cell(row=r, column=2, value=t).font = SMALL; r += 1

for name in wb.sheetnames:
    wb[name].sheet_view.showGridLines = False
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("gerado:", OUT)
