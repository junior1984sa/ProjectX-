# -*- coding: utf-8 -*-
"""Constroi a planilha de quantitativo de uma prancha de FORMAS.

Mesma identidade visual das planilhas de armadura, mas o quantitativo aqui e
de materiais (concreto, formas, grout, terraplenagem, impermeabilizacao e
drenagem), transcrito do quadro QUANTITATIVO TOTAL da prancha.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from estilo import (ACO, ALERTA, ALERTA_TXT, AZUL, BOX, CTR, CTRV, F_BODY, F_HDR,
                    F_SMALL, F_TOTAL, LARANJA, LEFT, LINHA, N_2, N_INT, N_PCT, TINTA,
                    TOP, VERDE, cabecalho, card, configurar_impressao, grafico_barras,
                    grafico_rosca, nota, secao, tabela, total_row)

# indices dos campos de um item do quantitativo
IT, UNI, QUN, QTO, GRP, OBS = range(6)


def gerar(spec, destino):
    nt = spec["nt"]
    itens = spec["itens"]
    fonte = spec["fonte"]
    grupos = spec["grupos"]

    def total_prancha(i):
        """Total impresso na prancha; quando ha uma unica unidade, e o proprio unitario."""
        return i[QTO] if i[QTO] is not None else i[QUN]

    vol = lambda i: i[UNI] == "m³"
    are = lambda i: i[UNI] == "m²"
    soma_grupo = lambda g, filtro: sum(total_prancha(i) for i in itens if i[GRP] == g and filtro(i))
    vol_total = sum(total_prancha(i) for i in itens if vol(i))
    area_total = sum(total_prancha(i) for i in itens if are(i))
    achar = lambda nome: next(total_prancha(i) for i in itens if i[IT] == nome)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard Executivo"
    ws_qt = wb.create_sheet("Quantitativo Total")
    ws_gr = wb.create_sheet("Resumo por Grupo")
    ws_un = wb.create_sheet("Quantitativo por Tanque")
    ws_geo = wb.create_sheet("Geometria e Seções")
    ws_cam = wb.create_sheet("Camadas e Materiais")
    ws_par = wb.create_sheet("Parâmetros Técnicos")
    ws_val = wb.create_sheet("Validações")
    ws_ler = wb.create_sheet("Leia-me")

    # ------------------------------------------------------------ Quantitativo Total
    cabecalho(ws_qt, "QUANTITATIVO TOTAL - TRANSCRIÇÃO INTEGRAL DA PRANCHA", fonte, 8)
    nota(ws_qt, 3, 8, spec["nota_qt"])
    dados = []
    for i in itens:
        tp = total_prancha(i)
        calc = i[QUN] * nt
        dados.append([i[IT], i[UNI], i[QUN], tp, round(calc, 2), round(tp - calc, 2),
                      i[GRP], i[OBS]])
    fim = tabela(ws_qt, 5,
        ["Item", "Unidade", "Quantidade\npor tanque", "Total impresso\nna prancha",
         "Total calculado\n(unit. × %d)" % nt, "Desvio", "Grupo", "Observação"],
        dados, [None, None, N_2, N_2, N_2, N_2, None, None],
        [40, 10, 13, 15, 15, 11, 26, 46])
    for i in range(6, fim + 1):
        if abs(ws_qt.cell(row=i, column=6).value or 0) > 0.05:
            for j in range(1, 9):
                ws_qt.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ALERTA)
                ws_qt.cell(row=i, column=j).font = Font(name="Arial", size=9, bold=True,
                                                        color=ALERTA_TXT)
    ws_qt.freeze_panes = "A6"

    # ------------------------------------------------------------ Resumo por Grupo
    cabecalho(ws_gr, "RESUMO POR GRUPO DE SERVIÇO", fonte, 6)
    nota(ws_gr, 3, 6,
         "Agrupamento dos itens do quadro por natureza de serviço. Volumes (m³), áreas (m²) e "
         "unidades são somados separadamente, pois não se agregam entre si.", altura=26)
    dados = []
    for g in grupos:
        sel = [i for i in itens if i[GRP] == g]
        dados.append([g, len(sel),
                      round(soma_grupo(g, vol), 2) or None,
                      round(soma_grupo(g, are), 2) or None,
                      round(sum(total_prancha(i) for i in sel if i[UNI] == "un"), 2) or None,
                      ", ".join(i[IT] for i in sel)])
    fim = tabela(ws_gr, 5,
        ["Grupo de serviço", "Itens", "Volume total\n(m³)", "Área total\n(m²)",
         "Unidades\n(un)", "Itens que compõem o grupo"],
        dados, [None, N_INT, N_2, N_2, N_INT, None], [30, 8, 14, 14, 11, 74])
    total_row(ws_gr, fim + 1, 6, "TOTAL",
              [len(itens), round(vol_total, 2), round(area_total, 2),
               sum(total_prancha(i) for i in itens if i[UNI] == "un"), None],
              [N_INT, N_2, N_2, N_INT, None], span=1)
    ws_gr.freeze_panes = "A6"

    # ------------------------------------------------------------ Quantitativo por Tanque
    cabecalho(ws_un, "QUANTITATIVO POR TANQUE", fonte, 5)
    nota(ws_un, 3, 5, spec["nota_unitario"])
    if nt > 1:
        chaves = spec["itens_por_tanque"]
        dados = [[tq, spec["elemento"]] + [achar(k) / nt for k in chaves] for tq in spec["tanques"]]
        cols = ["Tanque", "Elemento"] + [k for k in chaves]
        # a coluna por tanque usa o unitario impresso, nao o total dividido
        dados = [[tq, spec["elemento"]] +
                 [next(i[QUN] for i in itens if i[IT] == k) for k in chaves]
                 for tq in spec["tanques"]]
        fim = tabela(ws_un, 5, cols, dados, [None, None] + [N_2] * len(chaves),
                     [17, 26] + [15] * len(chaves))
        total_row(ws_un, fim + 1, len(cols), "TOTAL — %d tanques" % nt,
                  [round(achar(k), 2) for k in chaves], [N_2] * len(chaves), span=2)
        r = fim + 3
    else:
        r = 5
    secao(ws_un, r, 1, 4, "QUANTITATIVO COMPLETO DE UMA FUNDAÇÃO")
    uni = [[i[IT], i[UNI], i[QUN], i[GRP]] for i in itens]
    tabela(ws_un, r + 1, ["Item", "Unidade", "Quantidade por tanque", "Grupo"],
           uni, [None, None, N_2, None], [40, 10, 20, 26])
    ws_un.freeze_panes = "A6"

    # ------------------------------------------------------------ Geometria / Camadas / Parâmetros
    cabecalho(ws_geo, "GEOMETRIA, VISTAS E SEÇÕES", fonte, 6)
    nota(ws_geo, 3, 6,
         "Cotas, raios e elevações lidos diretamente da prancha, e grandezas derivadas a partir "
         "delas. Os itens marcados como DERIVADO não constam do desenho: são cálculo próprio.")
    tabela(ws_geo, 5, ["Grupo", "Item", "Escala", "Valor", "Origem", "Observação"],
           spec["geometria"], [None, None, None, N_2, None, None], [14, 42, 9, 14, 12, 66])
    ws_geo.freeze_panes = "A6"

    cabecalho(ws_cam, "CAMADAS, DETALHES E ESPECIFICAÇÃO DE MATERIAIS", fonte, 5)
    nota(ws_cam, 3, 5,
         "Composição das camadas do berço do tanque e dos detalhes construtivos, na ordem de "
         "execução, conforme o DETALHE 1 e as notas gerais da prancha.")
    tabela(ws_cam, 5, ["Ordem", "Camada / elemento", "Espessura ou cota", "Item do quantitativo",
                       "Especificação conforme a prancha"],
           spec["camadas"], [N_INT, None, None, None, None], [8, 34, 20, 30, 70])
    ws_cam.freeze_panes = "A6"

    cabecalho(ws_par, "PARÂMETROS TÉCNICOS E IDENTIFICAÇÃO DA PRANCHA", fonte, 3)
    nota(ws_par, 3, 3, "Dados de carimbo, notas gerais e documentos vinculados, transcritos da prancha.",
         altura=18)
    tabela(ws_par, 5, ["Grupo", "Parâmetro", "Especificação / conteúdo transcrito"],
           spec["parametros"], [None, None, None], [26, 36, 100])
    ws_par.freeze_panes = "A6"

    # ------------------------------------------------------------ Validações
    cabecalho(ws_val, "CONTROLES DE CONSISTÊNCIA", fonte, 6)
    nota(ws_val, 3, 6,
         "Conferência entre o quadro impresso e o produto quantidade unitária × número de tanques, "
         "além dos cruzamentos geométricos possíveis. Toda divergência está sinalizada.", altura=26)
    val = []
    if nt > 1:
        for i in itens:
            tp = total_prancha(i)
            calc = i[QUN] * nt
            dv = round(tp - calc, 2)
            ok = abs(dv) <= 0.05
            val.append(["Multiplicação" if ok else "DIVERGÊNCIA",
                        "%s — total vs. unitário × %d (%s)" % (i[IT], nt, i[UNI]),
                        round(calc, 2), tp, dv,
                        ("Total impresso confere com a quantidade unitária multiplicada pelos %d tanques."
                         % nt) if ok else
                        ("Total impresso equivale a %.4f × a quantidade unitária, e não a %d. "
                         "Diferença de %.2f %s. CONFIRMAR COM O PROJETISTA." %
                         (tp / i[QUN], nt, dv, i[UNI]))])
    val += spec["validacoes"]
    fim = tabela(ws_val, 5,
        ["Grupo", "Indicador", "Calculado", "Referência", "Desvio", "Critério / observação"],
        val, [None, None, N_2, N_2, N_2, None], [18, 46, 12, 12, 11, 92])
    for i in range(6, fim + 1):
        if ws_val.cell(row=i, column=1).value == "DIVERGÊNCIA":
            for j in range(1, 7):
                ws_val.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ALERTA)
                ws_val.cell(row=i, column=j).font = Font(name="Arial", size=9, bold=True,
                                                         color=ALERTA_TXT)
    ws_val.freeze_panes = "A6"

    cabecalho(ws_ler, "LEIA-ME - QUANTITATIVO DE FORMAS", fonte, 2)
    fim = tabela(ws_ler, 5, ["Tópico", "Conteúdo"], spec["leiame"], [None, None], [24, 150])
    for i in range(6, fim + 1):
        ws_ler.row_dimensions[i].height = 58
        ws_ler.cell(row=i, column=1).alignment = TOP
        ws_ler.cell(row=i, column=2).alignment = TOP

    # ------------------------------------------------------------ Dashboard
    ws = ws_dash
    cabecalho(ws, spec["titulo_painel"], spec["subtitulo_painel"], 14)
    for col, ln, tit, v, leg, cor, fmt in spec["cards"](achar, vol_total, area_total, nt, len(itens)):
        card(ws, col, ln, tit, v, leg, cor, fmt)
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[12].height = 16

    # volumes
    secao(ws, 15, 1, 5, "VOLUMES (m³) - TOTAL DA PRANCHA")
    ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=2)
    for j, h in ((1, "Item"), (3, "Volume (m³)"), (4, "% do volume"), (5, "Por tanque (m³)")):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=2).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=2).border = BOX
    ws.row_dimensions[16].height = 26
    lin = 17
    for i in [x for x in itens if vol(x)]:
        tp = total_prancha(i)
        ws.merge_cells(start_row=lin, start_column=1, end_row=lin, end_column=2)
        for j, v, f in ((1, i[IT], None), (3, tp, N_2), (4, tp / vol_total, N_PCT),
                        (5, i[QUN], N_2)):
            c = ws.cell(row=lin, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX
            c.alignment = LEFT if j == 1 else CTRV
            if f: c.number_format = f
        ws.cell(row=lin, column=2).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=lin, column=2).border = BOX
        lin += 1
    ws.merge_cells(start_row=lin, start_column=1, end_row=lin, end_column=2)
    for j, v, f in ((1, "TOTAL", None), (3, round(vol_total, 2), N_2), (4, 1.0, N_PCT),
                    (5, round(sum(i[QUN] for i in itens if vol(i)), 2), N_2)):
        c = ws.cell(row=lin, column=j, value=v)
        c.font = F_TOTAL; c.fill = PatternFill("solid", fgColor=ACO); c.border = BOX
        c.alignment = LEFT if j == 1 else CTRV
        if f: c.number_format = f
    ws.cell(row=lin, column=2).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=lin, column=2).border = BOX
    lin_vol = lin

    # grupos
    secao(ws, 15, 6, 9, "PARTICIPAÇÃO POR GRUPO DE SERVIÇO (VOLUME)")
    ws.merge_cells(start_row=16, start_column=6, end_row=16, end_column=7)
    for j, h in ((6, "Grupo de serviço"), (8, "Volume (m³)"), (9, "% do volume")):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=7).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=7).border = BOX
    lg = 17
    for g in grupos:
        sv = soma_grupo(g, vol)
        if not sv:
            continue
        ws.merge_cells(start_row=lg, start_column=6, end_row=lg, end_column=7)
        for j, v, f in ((6, g, None), (8, round(sv, 2), N_2), (9, sv / vol_total, N_PCT)):
            c = ws.cell(row=lg, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX
            c.alignment = LEFT if j == 6 else CTRV
            if f: c.number_format = f
        ws.cell(row=lg, column=7).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=lg, column=7).border = BOX
        lg += 1
    ws.merge_cells(start_row=lg, start_column=6, end_row=lg, end_column=7)
    for j, v, f in ((6, "TOTAL", None), (8, round(vol_total, 2), N_2), (9, 1.0, N_PCT)):
        c = ws.cell(row=lg, column=j, value=v)
        c.font = F_TOTAL; c.fill = PatternFill("solid", fgColor=ACO); c.border = BOX
        c.alignment = LEFT if j == 6 else CTRV
        if f: c.number_format = f
    ws.cell(row=lg, column=7).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=lg, column=7).border = BOX

    # areas
    secao(ws, 15, 11, 14, "ÁREAS (m²) - TOTAL DA PRANCHA")
    ws.merge_cells(start_row=16, start_column=11, end_row=16, end_column=12)
    for j, h in ((11, "Item"), (13, "Área (m²)"), (14, "Por tanque (m²)")):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=12).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=12).border = BOX
    la = 17
    for i in [x for x in itens if are(x)]:
        ws.merge_cells(start_row=la, start_column=11, end_row=la, end_column=12)
        for j, v, f in ((11, i[IT], None), (13, total_prancha(i), N_2), (14, i[QUN], N_2)):
            c = ws.cell(row=la, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX
            c.alignment = LEFT if j == 11 else CTRV
            if f: c.number_format = f
        ws.cell(row=la, column=12).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=la, column=12).border = BOX
        la += 1
    ws.merge_cells(start_row=la, start_column=11, end_row=la, end_column=12)
    for j, v, f in ((11, "TOTAL", None), (13, round(area_total, 2), N_2),
                    (14, round(sum(i[QUN] for i in itens if are(i)), 2), N_2)):
        c = ws.cell(row=la, column=j, value=v)
        c.font = F_TOTAL; c.fill = PatternFill("solid", fgColor=ACO); c.border = BOX
        c.alignment = LEFT if j == 11 else CTRV
        if f: c.number_format = f
    ws.cell(row=la, column=12).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=la, column=12).border = BOX

    r_graf = max(lin_vol, lg, la) + 2
    ws.add_chart(grafico_barras(ws, 3, 16, lin_vol - 1, 1, "Volume por item (m³)", "m³", "Item",
                                largura=19.0, altura=9.5), "A%d" % r_graf)
    ws.add_chart(grafico_rosca(ws, 8, 16, lg - 1, 6, "Participação por grupo (volume)",
                               largura=15.0, altura=9.5), "H%d" % r_graf)
    r_area = r_graf + 20
    secao(ws, r_area, 1, 14, "ÁREAS POR ITEM (m²) - TOTAL DA PRANCHA")
    ws.add_chart(grafico_barras(ws, 13, 16, la - 1, 11, "Área por item (m²)", "m²", "Item",
                                cor=LARANJA, largura=32.0, altura=8.4), "A%d" % (r_area + 2))
    r_dest = r_area + 20

    secao(ws, r_dest, 1, 14, "CONTROLE DE VALIDAÇÃO - DESTAQUES")
    ws.merge_cells("B%d:C%d" % (r_dest + 1, r_dest + 1))
    ws.merge_cells("G%d:N%d" % (r_dest + 1, r_dest + 1))
    for j, h in ((1, "Grupo"), (2, "Indicador"), (4, "Calculado"), (5, "Referência"),
                 (6, "Desvio"), (7, "Critério")):
        c = ws.cell(row=r_dest + 1, column=j, value=h)
        c.font = F_HDR; c.alignment = CTR
    for j in range(1, 15):
        ws.cell(row=r_dest + 1, column=j).fill = PatternFill("solid", fgColor=ACO)
        ws.cell(row=r_dest + 1, column=j).border = BOX
    for i, row in enumerate(spec["destaques"], start=r_dest + 2):
        alerta = row[0] == "DIVERGÊNCIA"
        fonte_l = (Font(name="Arial", size=8, bold=True, color=ALERTA_TXT) if alerta
                   else Font(name="Arial", size=8, color=TINTA))
        fundo = PatternFill("solid", fgColor=ALERTA if alerta else LINHA)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=14)
        for j, v in ((1, row[0]), (2, row[1]), (4, row[2]), (5, row[3]), (6, row[4]), (7, row[5])):
            c = ws.cell(row=i, column=j, value=v)
            c.font = fonte_l; c.fill = fundo; c.border = BOX
            c.alignment = LEFT if j in (2, 7) else CTRV
            if j in (4, 5, 6): c.number_format = N_2
        for j in range(1, 15):
            ws.cell(row=i, column=j).fill = fundo
            ws.cell(row=i, column=j).border = BOX
        ws.row_dimensions[i].height = 24
    for colL in "ABCDEFGHIJKLMN":
        ws.column_dimensions[colL].width = 13
    ws.freeze_panes = "A4"

    configurar_impressao(wb)
    wb.save(destino)
    return {"volume": vol_total, "area": area_total, "arquivo": destino}
