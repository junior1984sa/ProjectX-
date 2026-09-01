# -*- coding: utf-8 -*-
"""Constroi a planilha de analise de uma prancha de LAY-OUT / arranjo de equipamentos.

O construtor e neutro quanto ao idioma: todos os rotulos chegam prontos no
dicionario `spec["L"]`, o que permite gerar a versao inglesa e a portuguesa a
partir da mesma base de dados. A formatacao vem de `estilo.py`.
"""
import collections
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from estilo import (ACO, ALERTA, ALERTA_TXT, AZUL, BOX, CTR, CTRV, F_BODY, F_HDR,
                    F_SMALL, F_TOTAL, LARANJA, LEFT, LINHA, N_2, N_3, N_INT, N_PCT,
                    TINTA, TOP, VERDE, cabecalho, card, configurar_impressao,
                    grafico_barras, grafico_rosca, nota, secao, tabela, total_row)
import dados_u5700 as D


def gerar(spec, destino):
    L = spec["L"]
    idi = spec["idioma"]                       # "EN" ou "PT"
    desc = (lambda e: e[D.DEN]) if idi == "EN" else (lambda e: e[D.DPT])
    cat_nome = D.CAT_EN if idi == "EN" else D.CAT_PT
    grp_nome = D.GRP_EN if idi == "EN" else D.GRP_PT
    fonte = spec["fonte"]

    tags_por_cat = collections.Counter(e[D.CATG] for e in D.EQUIP)
    un_por_cat = collections.Counter()
    for e in D.EQUIP:
        un_por_cat[e[D.CATG]] += D.unidades(e[D.SUF])
    tags_por_grp = collections.Counter(e[D.GRP] for e in D.EQUIP)
    un_por_grp = collections.Counter()
    for e in D.EQUIP:
        un_por_grp[e[D.GRP]] += D.unidades(e[D.SUF])
    n_tags = len(D.EQUIP)
    n_un = sum(D.unidades(e[D.SUF]) for e in D.EQUIP)
    grupos = sorted(grp_nome)

    els = D.ELEVACOES
    el_min, el_max = els[0][0], els[-1][0]

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = L["ABA_DASH"]
    ws_lista = wb.create_sheet(L["ABA_LISTA"])
    ws_cat = wb.create_sheet(L["ABA_CAT"])
    ws_grp = wb.create_sheet(L["ABA_GRP"])
    ws_el = wb.create_sheet(L["ABA_EL"])
    ws_area = wb.create_sheet(L["ABA_AREA"])
    ws_site = wb.create_sheet(L["ABA_SITE"])
    ws_dwg = wb.create_sheet(L["ABA_DWG"])
    ws_val = wb.create_sheet(L["ABA_VAL"])
    ws_ler = wb.create_sheet(L["ABA_LER"])

    # ------------------------------------------------------------ lista de equipamentos
    cabecalho(ws_lista, L["T_LISTA"], fonte, 9)
    nota(ws_lista, 3, 9, L["N_LISTA"])
    dados = []
    for i, e in enumerate(D.EQUIP, start=1):
        dados.append([i, D.tag_cheia(e), e[D.TAG], e[D.SUF] or "-", D.unidades(e[D.SUF]),
                      cat_nome[e[D.CATG]], desc(e), grp_nome[e[D.GRP]],
                      L["IDI_PT"] if e[D.IDI] == "PT" else L["IDI_EN"]])
    fim = tabela(ws_lista, 5,
        [L["H_ITEM"], L["H_TAG"], L["H_TAGBASE"], L["H_SUF"], L["H_UN"], L["H_CAT"],
         L["H_DESC"], L["H_GRP"], L["H_IDI"]],
        dados, [N_INT, None, None, None, N_INT, None, None, None, None],
        [7, 12, 11, 10, 9, 24, 52, 30, 20])
    total_row(ws_lista, fim + 1, 9, L["TOTAL"],
              [None, None, n_un, "%d %s" % (len(D.CAT), L["W_CATS"]), None,
               "%d %s" % (len(grupos), L["W_GRPS"]), None],
              [None, None, N_INT, None, None, None, None], span=2)
    ws_lista.freeze_panes = "A6"

    # ------------------------------------------------------------ resumo por categoria
    cabecalho(ws_cat, L["T_CAT"], fonte, 6)
    nota(ws_cat, 3, 6, L["N_CAT"], altura=18)
    dados = [[cat_nome[c], D.CAT_PRANCHA[c], tags_por_cat[c], un_por_cat[c],
              un_por_cat[c] / n_un,
              ", ".join(D.tag_cheia(e) for e in D.EQUIP if e[D.CATG] == c)] for c in D.CAT]
    fim = tabela(ws_cat, 5,
        [L["H_CAT"], L["H_CATPR"], L["H_TAGS"], L["H_UN"], L["H_PCTUN"], L["H_TAGSLIST"]],
        dados, [None, None, N_INT, N_INT, N_PCT, None], [26, 22, 10, 11, 11, 86])
    total_row(ws_cat, fim + 1, 6, L["TOTAL"], [n_tags, n_un, 1.0, None],
              [N_INT, N_INT, N_PCT, None], span=2)
    ws_cat.freeze_panes = "A6"

    # ------------------------------------------------------------ grupos de processo
    cabecalho(ws_grp, L["T_GRP"], fonte, 5)
    nota(ws_grp, 3, 5, L["N_GRP"])
    dados = [[grp_nome[g], tags_por_grp[g], un_por_grp[g], tags_por_grp[g] / n_tags,
              ", ".join(D.tag_cheia(e) for e in D.EQUIP if e[D.GRP] == g)] for g in grupos]
    fim = tabela(ws_grp, 5,
        [L["H_GRP"], L["H_TAGS"], L["H_UN"], L["H_PCTTAG"], L["H_TAGSLIST"]],
        dados, [None, N_INT, N_INT, N_PCT, None], [34, 10, 11, 11, 96])
    total_row(ws_grp, fim + 1, 5, L["TOTAL"], [n_tags, n_un, 1.0, None],
              [N_INT, N_INT, N_PCT, None], span=1)
    ws_grp.freeze_panes = "A6"

    # ------------------------------------------------------------ elevações
    cabecalho(ws_el, L["T_EL"], fonte, 5)
    nota(ws_el, 3, 5, L["N_EL"])
    dados = [[i, v, v + D.CONV_NIVEL_MAR, n, round(v - el_min, 3)]
             for i, (v, n) in enumerate(els, start=1)]
    fim = tabela(ws_el, 5,
        [L["H_ORD"], L["H_ELPDMS"], L["H_ELMAR"], L["H_OCOR"], L["H_ACIMA"]],
        dados, [N_INT, N_3, N_3, N_INT, N_3], [8, 18, 20, 14, 20])
    total_row(ws_el, fim + 1, 5, L["TOTAL"],
              [None, None, sum(n for _, n in els), round(el_max - el_min, 3)],
              [None, None, N_INT, N_3], span=2)
    ws_el.freeze_panes = "A6"

    # ------------------------------------------------------------ áreas e coordenadas
    cabecalho(ws_area, L["T_AREA"], fonte, 5)
    nota(ws_area, 3, 5, L["N_AREA"])
    tabela(ws_area, 5, [L["H_GRUPO"], L["H_ITEMD"], L["H_VALOR"], L["H_UNID"], L["H_ORIG"]],
           spec["areas"], [None, None, N_3, None, None], [24, 52, 18, 12, 58])
    ws_area.freeze_panes = "A6"

    # ------------------------------------------------------------ entorno
    cabecalho(ws_site, L["T_SITE"], fonte, 3)
    nota(ws_site, 3, 3, L["N_SITE"], altura=18)
    tabela(ws_site, 5, [L["H_TIPO"], L["H_IDENT"], L["H_OBS"]],
           spec["entorno"], [None, None, None], [26, 34, 96])
    ws_site.freeze_panes = "A6"

    # ------------------------------------------------------------ dados da prancha
    cabecalho(ws_dwg, L["T_DWG"], fonte, 3)
    nota(ws_dwg, 3, 3, L["N_DWG"], altura=18)
    tabela(ws_dwg, 5, [L["H_GRUPO"], L["H_PARAM"], L["H_CONT"]],
           spec["prancha"], [None, None, None], [26, 34, 106])
    ws_dwg.freeze_panes = "A6"

    # ------------------------------------------------------------ validações
    cabecalho(ws_val, L["T_VAL"], fonte, 6)
    nota(ws_val, 3, 6, L["N_VAL"], altura=26)
    fim = tabela(ws_val, 5,
        [L["H_GRUPO"], L["H_IND"], L["H_CALC"], L["H_REF"], L["H_DESV"], L["H_CRIT"]],
        spec["validacoes"], [None, None, N_3, N_3, N_3, None], [20, 44, 14, 16, 12, 90])
    for i in range(6, fim + 1):
        if ws_val.cell(row=i, column=1).value == L["W_DIV"]:
            for j in range(1, 7):
                ws_val.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ALERTA)
                ws_val.cell(row=i, column=j).font = Font(name="Arial", size=9, bold=True,
                                                         color=ALERTA_TXT)
    ws_val.freeze_panes = "A6"

    # ------------------------------------------------------------ leia-me
    cabecalho(ws_ler, L["T_LER"], fonte, 2)
    fim = tabela(ws_ler, 5, [L["H_TOPICO"], L["H_CONT"]], spec["leiame"], [None, None], [24, 150])
    for i in range(6, fim + 1):
        ws_ler.row_dimensions[i].height = 58
        ws_ler.cell(row=i, column=1).alignment = TOP
        ws_ler.cell(row=i, column=2).alignment = TOP

    # ------------------------------------------------------------ painel
    ws = ws_dash
    cabecalho(ws, spec["titulo_painel"], spec["subtitulo_painel"], 14)
    cards = [
     (1, 4, L["C_TAGS"], n_tags, L["C_TAGS_L"], AZUL, N_INT),
     (4, 4, L["C_UN"], n_un, L["C_UN_L"], LARANJA, N_INT),
     (7, 4, L["C_AREA"], D.AREA_TOTAL, L["C_AREA_L"], ACO, N_INT),
     (10, 4, L["C_CAT"], len(D.CAT), L["C_CAT_L"], VERDE, N_INT),
     (1, 9, L["C_PUMP"], un_por_cat["PUMPS"], L["C_PUMP_L"], VERDE, N_INT),
     (4, 9, L["C_ENV"], round(el_max - el_min, 2), L["C_ENV_L"], ACO, N_2),
     (7, 9, L["C_TOP"], el_max, L["C_TOP_L"], LARANJA, N_3),
     (10, 9, L["C_GRP"], len(grupos), L["C_GRP_L"], AZUL, N_INT),
    ]
    for col, ln, tit, v, leg, cor, fmt in cards:
        card(ws, col, ln, tit, v, leg, cor, fmt)
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[12].height = 16

    # tabela de categorias (A:E)
    secao(ws, 15, 1, 5, L["S_CAT"])
    ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=2)
    for j, h in ((1, L["H_CAT"]), (3, L["H_TAGS"]), (4, L["H_UN"]), (5, L["H_PCTUN"])):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=2).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=2).border = BOX
    ws.row_dimensions[16].height = 26
    lin = 17
    for c_ in D.CAT:
        ws.merge_cells(start_row=lin, start_column=1, end_row=lin, end_column=2)
        for j, v, f in ((1, cat_nome[c_], None), (3, tags_por_cat[c_], N_INT),
                        (4, un_por_cat[c_], N_INT), (5, un_por_cat[c_] / n_un, N_PCT)):
            cc = ws.cell(row=lin, column=j, value=v)
            cc.font = F_BODY; cc.fill = PatternFill("solid", fgColor=LINHA); cc.border = BOX
            cc.alignment = LEFT if j == 1 else CTRV
            if f: cc.number_format = f
        ws.cell(row=lin, column=2).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=lin, column=2).border = BOX
        lin += 1
    ws.merge_cells(start_row=lin, start_column=1, end_row=lin, end_column=2)
    for j, v, f in ((1, L["TOTAL"], None), (3, n_tags, N_INT), (4, n_un, N_INT), (5, 1.0, N_PCT)):
        cc = ws.cell(row=lin, column=j, value=v)
        cc.font = F_TOTAL; cc.fill = PatternFill("solid", fgColor=ACO); cc.border = BOX
        cc.alignment = LEFT if j == 1 else CTRV
        if f: cc.number_format = f
    ws.cell(row=lin, column=2).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=lin, column=2).border = BOX
    lin_cat = lin

    # tabela de grupos (H:K)
    secao(ws, 15, 8, 11, L["S_GRP"])
    ws.merge_cells(start_row=16, start_column=8, end_row=16, end_column=9)
    for j, h in ((8, L["H_GRP"]), (10, L["H_TAGS"]), (11, L["H_PCTTAG"])):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=9).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=9).border = BOX
    lg = 17
    for g in grupos:
        ws.merge_cells(start_row=lg, start_column=8, end_row=lg, end_column=9)
        for j, v, f in ((8, grp_nome[g], None), (10, tags_por_grp[g], N_INT),
                        (11, tags_por_grp[g] / n_tags, N_PCT)):
            cc = ws.cell(row=lg, column=j, value=v)
            cc.font = F_BODY; cc.fill = PatternFill("solid", fgColor=LINHA); cc.border = BOX
            cc.alignment = LEFT if j == 8 else CTRV
            if f: cc.number_format = f
        ws.cell(row=lg, column=9).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=lg, column=9).border = BOX
        lg += 1
    ws.merge_cells(start_row=lg, start_column=8, end_row=lg, end_column=9)
    for j, v, f in ((8, L["TOTAL"], None), (10, n_tags, N_INT), (11, 1.0, N_PCT)):
        cc = ws.cell(row=lg, column=j, value=v)
        cc.font = F_TOTAL; cc.fill = PatternFill("solid", fgColor=ACO); cc.border = BOX
        cc.alignment = LEFT if j == 8 else CTRV
        if f: cc.number_format = f
    ws.cell(row=lg, column=9).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=lg, column=9).border = BOX

    r_graf = max(lin_cat, lg) + 2
    ws.add_chart(grafico_barras(ws, 4, 16, lin_cat - 1, 1, L["G_UN"], L["G_UN_Y"], L["G_UN_X"],
                                largura=19.0, altura=9.5), "A%d" % r_graf)
    ws.add_chart(grafico_rosca(ws, 10, 16, lg - 1, 8, L["G_GRP"], largura=15.0, altura=9.5),
                 "H%d" % r_graf)

    r_el = r_graf + 20
    secao(ws, r_el, 1, 14, L["S_EL"])
    ws.add_chart(grafico_barras(ws_el, 4, 5, 5 + len(els), 2, L["G_EL"], L["G_EL_Y"], L["G_EL_X"],
                                cor=ACO, largura=32.0, altura=8.6), "A%d" % (r_el + 2))

    r_dest = r_el + 20
    secao(ws, r_dest, 1, 14, L["S_DEST"])
    ws.merge_cells("B%d:C%d" % (r_dest + 1, r_dest + 1))
    ws.merge_cells("G%d:N%d" % (r_dest + 1, r_dest + 1))
    for j, h in ((1, L["H_GRUPO"]), (2, L["H_IND"]), (4, L["H_CALC"]), (5, L["H_REF"]),
                 (6, L["H_DESV"]), (7, L["H_CRIT"])):
        c = ws.cell(row=r_dest + 1, column=j, value=h)
        c.font = F_HDR; c.alignment = CTR
    for j in range(1, 15):
        ws.cell(row=r_dest + 1, column=j).fill = PatternFill("solid", fgColor=ACO)
        ws.cell(row=r_dest + 1, column=j).border = BOX
    for i, row in enumerate(spec["destaques"], start=r_dest + 2):
        alerta = row[0] == L["W_DIV"]
        fonte_l = (Font(name="Arial", size=8, bold=True, color=ALERTA_TXT) if alerta
                   else Font(name="Arial", size=8, color=TINTA))
        fundo = PatternFill("solid", fgColor=ALERTA if alerta else LINHA)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=14)
        for j, v in ((1, row[0]), (2, row[1]), (4, row[2]), (5, row[3]), (6, row[4]), (7, row[5])):
            c = ws.cell(row=i, column=j, value=v)
            c.font = fonte_l; c.fill = fundo; c.border = BOX
            c.alignment = LEFT if j in (2, 7) else CTRV
            if j in (4, 5, 6) and isinstance(v, (int, float)): c.number_format = N_3
        for j in range(1, 15):
            ws.cell(row=i, column=j).fill = fundo
            ws.cell(row=i, column=j).border = BOX
        ws.row_dimensions[i].height = 24
    for colL in "ABCDEFGHIJKLMN":
        ws.column_dimensions[colL].width = 13
    ws.freeze_panes = "A4"

    configurar_impressao(wb, aba_pagina_unica=L["ABA_DASH"])
    wb.save(destino)
    return {"tags": n_tags, "unidades": n_un, "arquivo": destino}
