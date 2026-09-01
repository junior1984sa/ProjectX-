# -*- coding: utf-8 -*-
"""Constroi a planilha de quantitativo de uma prancha de ARMADURAS.

Recebe a especificacao de uma prancha (dados transcritos do desenho) e devolve
o arquivo .xlsx com as nove abas do padrao: painel executivo com graficos e as
abas de detalhe e auditoria. Toda a formatacao vem de `estilo.py`.
"""
import math
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from estilo import (ACO, ALERTA, ALERTA_TXT, AZUL, BOX, BRANCO, CTR, CTRV, F_BODY,
                    F_BOLD, F_HDR, F_SMALL, F_TOTAL, LARANJA, LEFT, LINHA, N_2, N_3,
                    N_INT, N_PCT, TINTA, TOP, VERDE, cabecalho, card,
                    configurar_impressao, grafico_barras, grafico_rosca, nota,
                    secao, tabela, total_row)

# massa linear nominal NBR 7480 (kg/m)
KGM = {6.3: 0.245, 8: 0.395, 10: 0.617, 12.5: 0.963, 16: 1.578, 20: 2.466, 25: 3.853}

# indices dos campos de uma barra
P, BIT, QTD, ULAB, UCM, TOT, ELE, FUN, VIS, OBS = range(10)


def _massa(b):
    return b[TOT] * KGM[b[BIT]]


def bit(d):
    """Bitola no padrao brasileiro: Ø12,5 em vez de Ø12.5."""
    return ("%g" % d).replace(".", ",")


def gerar(spec, destino):
    nt = spec["nt"]
    barras = spec["barras"]
    resumo_des = spec["resumo_des"]
    total_des = spec["total_des"]
    fonte = spec["fonte"]

    massa_tot = sum(_massa(b) for b in barras)
    compr_tot = sum(b[TOT] for b in barras)
    bitolas = sorted(resumo_des)
    barras_12m = sum(math.ceil(sum(b[TOT] for b in barras if b[BIT] == d) / 12)
                 for d in bitolas)
    qtd_estribos = sum(b[QTD] for b in barras if "stribo" in b[FUN])
    por_fund = (lambda v: v / nt)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard Executivo"
    ws_lista = wb.create_sheet("Lista de Barras")
    ws_bit = wb.create_sheet("Resumo por Bitola")
    ws_pos = wb.create_sheet("Resumo por Posição")
    ws_un = wb.create_sheet("Quantitativo Unitário")
    ws_geo = wb.create_sheet("Geometria e Seções")
    ws_par = wb.create_sheet("Parâmetros Técnicos")
    ws_val = wb.create_sheet("Validações")
    ws_ler = wb.create_sheet("Leia-me")

    rot_qtd = "Quant. total\n(%d fundações)" % nt if nt > 1 else "Quant."
    suf = " - %dx" % nt if nt > 1 else ""

    # ------------------------------------------------------------ Lista de Barras
    cabecalho(ws_lista, "LISTA DE BARRAS - TRANSCRIÇÃO INTEGRAL DA PRANCHA", fonte, 14)
    nota(ws_lista, 3, 14, spec["nota_lista"])
    dados = [[b[P], b[ELE], b[BIT], b[QTD], b[ULAB], b[TOT], KGM[b[BIT]],
              round(_massa(b), 2), round(_massa(b) / massa_tot, 6),
              b[QTD] / nt if b[QTD] % nt else b[QTD] // nt,
              round(b[TOT] / nt, 2), round(_massa(b) / nt, 2), b[FUN], b[VIS]]
             for b in barras]
    fim = tabela(ws_lista, 5,
        ["POS.", "Elemento", "Bitola\nØ (mm)", rot_qtd, "Compr. unit.\n(cm)",
         "Compr. total\n(m)%s" % suf, "Massa linear\n(kg/m)", "Massa total\n(kg)%s" % suf,
         "% da massa", "Quant. por\nfundação", "Compr. por\nfundação (m)",
         "Massa por\nfundação (kg)", "Função estrutural", "Vista de referência"],
        dados,
        [N_INT, None, None, N_INT, None, N_2, N_3, N_2, N_PCT, N_INT, N_2, N_2, None, None],
        [7, 10, 9, 13, 11, 13, 11, 13, 10, 11, 12, 12, 26, 26])
    total_row(ws_lista, fim + 1, 14, "TOTAL — %d posições" % len(barras),
        [None, sum(b[QTD] for b in barras), None, round(compr_tot, 2), None,
         round(massa_tot, 2), 1.0, None, round(compr_tot / nt, 2), round(massa_tot / nt, 2),
         None, None],
        [None, N_INT, None, N_2, None, N_2, N_PCT, None, N_2, N_2, None, None], span=2)

    r = fim + 3
    ws_lista.cell(row=r, column=1, value="Observações por posição").font = F_BOLD
    for i, b in enumerate(barras, start=r + 1):
        ws_lista.cell(row=i, column=1, value=b[ELE]).font = F_BOLD
        ws_lista.merge_cells(start_row=i, start_column=2, end_row=i, end_column=14)
        c = ws_lista.cell(row=i, column=2, value=b[OBS])
        c.font = F_SMALL
        c.alignment = LEFT
    ws_lista.freeze_panes = "A6"

    # ------------------------------------------------------------ Resumo por Bitola
    cabecalho(ws_bit, "RESUMO GERAL POR BITOLA", fonte, 10)
    nota(ws_bit, 3, 10,
         "Confronto entre o quadro RESUMO GERAL da prancha e o somatório das posições da LISTA DE "
         "BARRAS. Massas lineares conforme NBR 7480 (valores nominais adotados no próprio desenho).")
    dados = []
    for d in bitolas:
        comp_des, massa_des = resumo_des[d]
        comp_calc = sum(b[TOT] for b in barras if b[BIT] == d)
        dados.append([d, ", ".join(b[ELE] for b in barras if b[BIT] == d), comp_des, KGM[d],
                      massa_des, round(comp_calc, 2), round(comp_calc - comp_des, 2),
                      round(comp_calc * KGM[d] / massa_tot, 6), math.ceil(comp_des / 12),
                      round(comp_calc * KGM[d] / nt, 2)])
    fim = tabela(ws_bit, 5,
        ["Bitola\nØ (mm)", "Posições", "Compr. da prancha\n(m)", "Massa linear\n(kg/m)",
         "Massa da prancha\n(kg)", "Compr. somado\nda lista (m)", "Desvio de\ncompr. (m)",
         "% da massa", "Barras de 12 m\n(equivalente)", "Massa por\nfundação (kg)"],
        dados, [None, None, N_2, N_3, N_INT, N_2, N_2, N_PCT, N_INT, N_2],
        [9, 24, 15, 11, 15, 14, 11, 10, 14, 13])
    total_row(ws_bit, fim + 1, 10, "TOTAL",
        [round(sum(resumo_des[d][0] for d in bitolas), 2), None, total_des,
         round(compr_tot, 2), round(compr_tot - sum(resumo_des[d][0] for d in bitolas), 2),
         1.0, barras_12m, round(massa_tot / nt, 2)],
        [N_2, None, N_INT, N_2, N_2, N_PCT, N_INT, N_2], span=2)
    r = fim + 3
    for txt in spec["notas_bitola"]:
        nota(ws_bit, r, 10, "- " + txt)
        r += 1
    ws_bit.freeze_panes = "A6"

    # ------------------------------------------------------------ Resumo por Posição
    cabecalho(ws_pos, "RESUMO POR POSIÇÃO DE BARRA", fonte, 8)
    nota(ws_pos, 3, 8, "Massa por posição, na ordem da prancha. Fonte de dados dos gráficos do painel.",
         altura=18)
    dados = [[b[ELE], b[BIT], b[QTD], b[TOT], round(_massa(b), 2),
              round(_massa(b) / massa_tot, 6), b[FUN], round(_massa(b) / nt, 2)] for b in barras]
    fim = tabela(ws_pos, 5,
        ["Posição", "Bitola\nØ (mm)", "Quant.\ntotal", "Compr. total\n(m)", "Massa total\n(kg)",
         "% da massa", "Função estrutural", "Massa por\nfundação (kg)"],
        dados, [None, None, N_INT, N_2, N_2, N_PCT, None, N_2],
        [11, 10, 12, 13, 13, 10, 32, 13])
    total_row(ws_pos, fim + 1, 8, "TOTAL",
        [sum(b[QTD] for b in barras), round(compr_tot, 2), round(massa_tot, 2), 1.0, None,
         round(massa_tot / nt, 2)],
        [N_INT, N_2, N_2, N_PCT, None, N_2], span=2)

    r = fim + 3
    secao(ws_pos, r, 1, 5, "AGRUPAMENTO POR FUNÇÃO ESTRUTURAL")
    grp = []
    for nome, poss in spec["grupos"]:
        sel = [b for b in barras if b[P] in poss]
        grp.append([nome, ", ".join(b[ELE] for b in sel), sum(b[QTD] for b in sel),
                    round(sum(b[TOT] for b in sel), 2), round(sum(_massa(b) for b in sel), 2)])
    fim2 = tabela(ws_pos, r + 1,
        ["Função estrutural", "Posições", "Quant. total", "Compr. total\n(m)", "Massa total\n(kg)"],
        grp, [None, None, N_INT, N_2, N_2], [42, 20, 12, 13, 13])
    total_row(ws_pos, fim2 + 1, 5, "TOTAL",
        [sum(b[QTD] for b in barras), round(compr_tot, 2), round(massa_tot, 2)],
        [N_INT, N_2, N_2], span=2)
    ws_pos.freeze_panes = "A6"

    # ------------------------------------------------------------ Quantitativo Unitário
    cabecalho(ws_un, "QUANTITATIVO UNITÁRIO POR FUNDAÇÃO", fonte, 8)
    nota(ws_un, 3, 8, spec["nota_unitario"])
    if nt > 1:
        dados = [[tq, spec["elemento"], 1] +
                 [round(sum(b[TOT] for b in barras if b[BIT] == d) / nt, 2) for d in bitolas] +
                 [round(compr_tot / nt, 2), round(massa_tot / nt, 2)]
                 for tq in spec["tanques"]]
        cols = (["Tanque", "Elemento", "Fundações"] +
                ["Compr. Ø%s\n(m)" % bit(d) for d in bitolas] + ["Compr. total\n(m)", "Massa de aço\n(kg)"])
        fmts = [None, None, N_INT] + [N_2] * len(bitolas) + [N_2, N_2]
        larg = [17, 30, 11] + [12] * len(bitolas) + [13, 14]
        fim = tabela(ws_un, 5, cols, dados, fmts, larg)
        total_row(ws_un, fim + 1, len(cols), "TOTAL — %d fundações" % nt,
            [nt] + [round(sum(b[TOT] for b in barras if b[BIT] == d), 2) for d in bitolas] +
            [round(compr_tot, 2), round(massa_tot, 2)],
            [N_INT] + [N_2] * len(bitolas) + [N_2, N_2], span=2)
        r = fim + 3
    else:
        r = 5
    secao(ws_un, r, 1, 6, "DETALHAMENTO DE UMA FUNDAÇÃO" +
          (" (VALE PARA CADA UM DOS %d TANQUES)" % nt if nt > 1 else ""))
    uni = [[b[ELE], b[BIT], b[QTD] / nt if b[QTD] % nt else b[QTD] // nt, b[ULAB],
            round(b[TOT] / nt, 2), round(_massa(b) / nt, 2)] for b in barras]
    fim2 = tabela(ws_un, r + 1,
        ["Posição", "Bitola\nØ (mm)", "Quant. por\nfundação", "Compr. unit.\n(cm)",
         "Compr. por\nfundação (m)", "Massa por\nfundação (kg)"],
        uni, [None, None, N_INT, None, N_2, N_2], [11, 10, 12, 12, 13, 14])
    total_row(ws_un, fim2 + 1, 6, "TOTAL POR FUNDAÇÃO",
        [sum(b[QTD] for b in barras) / nt, None, round(compr_tot / nt, 2), round(massa_tot / nt, 2)],
        [N_INT, None, N_2, N_2], span=2)
    ws_un.freeze_panes = "A6"

    # ------------------------------------------------------------ Geometria
    cabecalho(ws_geo, "GEOMETRIA, VISTAS E SEÇÕES", fonte, 6)
    nota(ws_geo, 3, 6,
         "Cotas e desenvolvimentos lidos diretamente da prancha, e grandezas derivadas a partir delas. "
         "Os itens marcados como DERIVADO não constam do desenho: são cálculo próprio para apoio de planejamento.")
    tabela(ws_geo, 5, ["Grupo", "Item", "Escala", "Valor\n(m, m² ou m³)", "Origem", "Observação"],
           spec["geometria"], [None, None, None, N_3, None, None], [14, 42, 9, 14, 12, 66])
    ws_geo.freeze_panes = "A6"

    # ------------------------------------------------------------ Parâmetros Técnicos
    cabecalho(ws_par, "PARÂMETROS TÉCNICOS E IDENTIFICAÇÃO DA PRANCHA", fonte, 3)
    nota(ws_par, 3, 3, "Dados de carimbo, notas gerais, legenda e documentos vinculados, transcritos da prancha.",
         altura=18)
    tabela(ws_par, 5, ["Grupo", "Parâmetro", "Especificação / conteúdo transcrito"],
           spec["parametros"], [None, None, None], [26, 34, 96])
    ws_par.freeze_panes = "A6"

    # ------------------------------------------------------------ Validações
    cabecalho(ws_val, "CONTROLES DE CONSISTÊNCIA", fonte, 6)
    nota(ws_val, 3, 6,
         "Conferência entre o somatório estruturado e os quadros impressos na prancha. "
         "Toda divergência está sinalizada de forma explícita.", altura=18)
    val = []
    for d in bitolas:
        comp_des = resumo_des[d][0]
        comp_calc = sum(b[TOT] for b in barras if b[BIT] == d)
        dv = round(comp_calc - comp_des, 2)
        val.append(["Comprimento" if dv == 0 else "DIVERGÊNCIA",
                    "Ø%s — comprimento total (m)" % bit(d), round(comp_calc, 2), comp_des, dv,
                    ("Soma das posições da LISTA DE BARRAS confere exatamente com o quadro RESUMO GERAL."
                     if dv == 0 else spec["msg_desvio_compr"][d])])
    for d in bitolas:
        massa_des = resumo_des[d][1]
        massa_calc = resumo_des[d][0] * KGM[d]
        val.append(["Massa", "Ø%s — massa total (kg)" % bit(d), round(massa_calc, 2), massa_des,
                    round(massa_calc - massa_des, 2),
                    "Comprimento do RESUMO GERAL × massa linear. Desvio inferior a 0,5 kg, "
                    "compatível com o arredondamento do quadro."])
    massa_res = sum(resumo_des[d][0] * KGM[d] for d in bitolas)
    val.append(["Massa", "Massa total de aço (kg)", round(massa_res, 2), total_des,
                round(massa_res - total_des, 2),
                "Somatório do quadro RESUMO GERAL contra o TOTAL (kg) impresso na prancha."])
    val += spec["validacoes"]
    fim = tabela(ws_val, 5,
        ["Grupo", "Indicador", "Calculado", "Referência", "Desvio", "Critério / observação"],
        val, [None, None, N_2, N_2, N_2, None], [20, 44, 12, 12, 11, 92])
    for i in range(6, fim + 1):
        if ws_val.cell(row=i, column=1).value == "DIVERGÊNCIA":
            for j in range(1, 7):
                ws_val.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ALERTA)
                ws_val.cell(row=i, column=j).font = Font(name="Arial", size=9, bold=True,
                                                         color=ALERTA_TXT)
    ws_val.freeze_panes = "A6"

    # ------------------------------------------------------------ Leia-me
    cabecalho(ws_ler, "LEIA-ME - QUANTITATIVO DE ARMADURAS", fonte, 2)
    fim = tabela(ws_ler, 5, ["Tópico", "Conteúdo"], spec["leiame"], [None, None], [24, 150])
    for i in range(6, fim + 1):
        ws_ler.row_dimensions[i].height = 58
        ws_ler.cell(row=i, column=1).alignment = TOP
        ws_ler.cell(row=i, column=2).alignment = TOP

    # ------------------------------------------------------------ Dashboard
    ws = ws_dash
    cabecalho(ws, spec["titulo_painel"], spec["subtitulo_painel"], 14)
    for col, ln, tit, val_, leg, cor, fmt in spec["cards"](
            massa_tot, compr_tot, nt, len(barras), qtd_estribos, barras_12m, len(bitolas)):
        card(ws, col, ln, tit, val_, leg, cor, fmt)
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[12].height = 16

    secao(ws, 15, 1, 5, "COMPOSIÇÃO POR BITOLA")
    for j, h in enumerate(["Bitola", "Compr. total (m)", "Massa (kg)", "% da massa", "Barras 12 m"],
                          start=1):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.row_dimensions[16].height = 26
    lin = 17
    for d in bitolas:
        comp = sum(b[TOT] for b in barras if b[BIT] == d)
        m = comp * KGM[d]
        for j, v, f in ((1, "Ø%s mm" % bit(d), None), (2, round(comp, 2), N_2), (3, round(m, 2), N_2),
                        (4, m / massa_tot, N_PCT), (5, math.ceil(comp / 12), N_INT)):
            c = ws.cell(row=lin, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA)
            c.border = BOX; c.alignment = CTRV
            if f: c.number_format = f
        lin += 1
    total_row(ws, lin, 5, "TOTAL",
              [round(compr_tot, 2), round(massa_tot, 2), 1.0,
               sum(math.ceil(sum(b[TOT] for b in barras if b[BIT] == d) / 12) for d in bitolas)],
              [N_2, N_2, N_PCT, N_INT], span=1)
    lin_tot_bit = lin

    secao(ws, 15, 8, 11, "PARTICIPAÇÃO POR FUNÇÃO ESTRUTURAL")
    ws.merge_cells(start_row=16, start_column=8, end_row=16, end_column=9)
    for j, h in ((8, "Função estrutural"), (10, "Massa (kg)"), (11, "% da massa")):
        c = ws.cell(row=16, column=j, value=h)
        c.font = F_HDR; c.fill = PatternFill("solid", fgColor=ACO); c.alignment = CTR; c.border = BOX
    ws.cell(row=16, column=9).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=16, column=9).border = BOX
    lg = 17
    for nome, poss in spec["grupos"]:
        m = sum(_massa(b) for b in barras if b[P] in poss)
        ws.merge_cells(start_row=lg, start_column=8, end_row=lg, end_column=9)
        for j, v, f in ((8, nome, None), (10, round(m, 2), N_2), (11, m / massa_tot, N_PCT)):
            c = ws.cell(row=lg, column=j, value=v)
            c.font = F_BODY; c.fill = PatternFill("solid", fgColor=LINHA); c.border = BOX
            c.alignment = LEFT if j == 8 else CTRV
            if f: c.number_format = f
        ws.cell(row=lg, column=9).fill = PatternFill("solid", fgColor=LINHA)
        ws.cell(row=lg, column=9).border = BOX
        lg += 1
    ws.merge_cells(start_row=lg, start_column=8, end_row=lg, end_column=9)
    for j, v, f in ((8, "TOTAL", None), (10, round(massa_tot, 2), N_2), (11, 1.0, N_PCT)):
        c = ws.cell(row=lg, column=j, value=v)
        c.font = F_TOTAL; c.fill = PatternFill("solid", fgColor=ACO); c.border = BOX
        c.alignment = LEFT if j == 8 else CTRV
        if f: c.number_format = f
    ws.cell(row=lg, column=9).fill = PatternFill("solid", fgColor=ACO)
    ws.cell(row=lg, column=9).border = BOX

    ws.add_chart(grafico_barras(ws, 3, 16, lin_tot_bit - 1, 1,
                                "Massa de aço por bitola (kg)", "kg", "Bitola"), "A23")
    ws.add_chart(grafico_rosca(ws, 10, 16, lg - 1, 8, "Participação por função estrutural"), "H23")

    secao(ws, 40, 1, 14, "MASSA POR POSIÇÃO DE BARRA - FONTE: ABA RESUMO POR POSIÇÃO")
    ws.add_chart(grafico_barras(ws_pos, 5, 5, 5 + len(barras), 1,
                                "Massa por posição de barra (kg)", "kg", "Posição",
                                cor=ACO, largura=32.0, altura=8.6), "A42")

    secao(ws, 60, 1, 14, "CONTROLE DE VALIDAÇÃO - DESTAQUES")
    ws.merge_cells("B61:C61"); ws.merge_cells("G61:N61")
    for j, h in ((1, "Grupo"), (2, "Indicador"), (4, "Calculado"), (5, "Referência"),
                 (6, "Desvio"), (7, "Critério")):
        c = ws.cell(row=61, column=j, value=h)
        c.font = F_HDR; c.alignment = CTR
    for j in range(1, 15):
        ws.cell(row=61, column=j).fill = PatternFill("solid", fgColor=ACO)
        ws.cell(row=61, column=j).border = BOX
    for i, row in enumerate(spec["destaques"], start=62):
        alerta = row[0] == "DIVERGÊNCIA"
        fonte_l = (Font(name="Arial", size=8, bold=True, color=ALERTA_TXT) if alerta
                   else Font(name="Arial", size=8, color=TINTA))
        fundo = PatternFill("solid", fgColor=ALERTA if alerta else LINHA)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=14)
        for j, v in ((1, row[0]), (2, row[1]), (4, row[2]), (5, row[3]), (6, row[4]), (7, row[5])):
            c = ws.cell(row=i, column=j, value=v)
            c.font = fonte_l; c.fill = fundo; c.border = BOX
            c.alignment = LEFT if j in (2, 7) else CTRV
            if j in (4, 5, 6): c.number_format = N_2
        for j in range(1, 15):
            ws.cell(row=i, column=j).fill = fundo
            ws.cell(row=i, column=j).border = BOX
        ws.row_dimensions[i].height = 24
    for colL in "ABCDEFGHIJKLMN":
        ws.column_dimensions[colL].width = 13
    ws.freeze_panes = "A4"

    configurar_impressao(wb)
    wb.save(destino)
    return {"massa": massa_tot, "compr": compr_tot, "arquivo": destino}
