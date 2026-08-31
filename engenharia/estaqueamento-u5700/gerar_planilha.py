# -*- coding: utf-8 -*-
"""
Gera a planilha de controle de estaqueamento da U-5700 a partir do dados.json
produzido por extrair_coordenadas.py.

Uso:  python3 gerar_planilha.py [dados.json] [saida.xlsx]

Abas: Resumo + Casa dos compressores + Fornos + Torre + Reatores.
Cada aba traz a tabela de coordenadas (PDMS e GLOBAL SIRGAS 2000) e, por tag,
o controle das 4 etapas: terraplanagem, estaqueamento, pit e arrasamento.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ENTRADA = sys.argv[1] if len(sys.argv) > 1 else 'dados.json'
SAIDA = sys.argv[2] if len(sys.argv) > 2 else 'Estaqueamento_U-5700_Coordenadas.xlsx'
D = json.load(open(ENTRADA))
DATUM_E, DATUM_N = 714955.434, 7488954.390
ABAS = [('Casa dos compressores', 'Casa dos Compressores'),
        ('Fornos', 'Fornos'), ('Torre', 'Torre'), ('Reatores', 'Reatores')]
STATUS = '"NÃO INICIADO,EM ANDAMENTO,CONCLUÍDO,N/A"'

F   = 'Arial'
NAVY   = '1F3864'; BLUE = '2E5C8A'; LBLUE = 'D9E2F3'; GREY = 'F2F2F2'
YELLOW = 'FFF2CC'; GREEN = 'C6EFCE'; GREENT = '006100'
AMBER  = 'FFEB9C'; AMBERT = '9C6500'; REDF = 'FFC7CE'; REDT = '9C0006'

thin  = Side(style='thin',   color='BFBFBF')
med   = Side(style='medium', color='808080')
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(c, txt, fill=BLUE, size=9, rot=0):
    c.value = txt
    c.font = Font(name=F, size=size, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=fill)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, textRotation=rot)
    c.border = Border(left=med, right=med, top=med, bottom=med)

wb = Workbook()

# ───────────────────────────── RESUMO ─────────────────────────────
ws = wb.active; ws.title = 'Resumo'
ws.sheet_view.showGridLines = False
def put(r, c, v, bold=False, size=10, color='000000', fill=None, al='left', fmt=None, wrap=False):
    cel = ws.cell(row=r, column=c, value=v)
    cel.font = Font(name=F, size=size, bold=bold, color=color)
    cel.alignment = Alignment(horizontal=al, vertical='center', wrap_text=wrap)
    if fill: cel.fill = PatternFill('solid', fgColor=fill)
    if fmt:  cel.number_format = fmt
    return cel

ws.merge_cells('A1:I1')
put(1, 1, 'CONTROLE DE ESTAQUEAMENTO — U-5700 UNIDADE DE HIDROISODESPARAFINAÇÃO', True, 14, 'FFFFFF', NAVY, 'center')
ws.row_dimensions[1].height = 30
ws.merge_cells('A2:I2')
put(2, 1, 'REFINO BOAVENTURA — TR / BOAVENTURA — CONSÓRCIO TEM BOAVENTURA — Planilha de coordenadas de locação e controle de execução das estacas', False, 10, '444444', LBLUE, 'center')

put(4, 1, 'SISTEMA DE COORDENADAS', True, 11, NAVY)
put(5, 1, 'Sistema PDMS (E3D/NAVIS) — equivalência com o Sistema Geodésico Brasileiro, Datum horizontal SIRGAS 2000 (nota geral dos desenhos):')
put(6, 1, 'E (PDMS) = 0 m  →  E (GLOBAL) ='); put(6, 3, DATUM_E, True, 10, '0000FF', YELLOW, 'right', '#,##0.000')
put(7, 1, 'N (PDMS) = 0 m  →  N (GLOBAL) ='); put(7, 3, DATUM_N, True, 10, '0000FF', YELLOW, 'right', '#,##0.000')
put(8, 1, 'Elevação 0,00 (PDMS) = EL. 26,14 (nível do mar) — marégrafo de Imbituba/SC.')
put(9, 1, 'Dimensões em centímetro, elevações e coordenadas em metro, exceto onde indicado.')

# --- inventário ---
put(11, 1, 'INVENTÁRIO DE ESTACAS POR ABA', True, 11, NAVY)
inv_h = ['ABA', 'ESTRUTURA', 'FAIXA DE TAGS', 'Ø', 'QTD.', 'COTA DE ARRAS.', 'COMPR. PREV. (m)', 'DESENHO DE REFERÊNCIA', 'ORIGEM DAS COORDENADAS']
for j, h in enumerate(inv_h, 1): hdr(ws.cell(row=12, column=j), h)
ws.row_dimensions[12].height = 30
r = 13
for aba, k in ABAS:
    v = D[k]
    for g in v['grupos']:
        vals = [aba, g['estrutura'], g['faixa'], g['diam'], g['qtd'], g['arras'], g['lest'],
                f"{v['doc']} rev.{v['rev']} ({v['data']})", v['fonte']]
        for j, x in enumerate(vals, 1):
            put(r, j, x, al='center' if j in (4, 5, 6, 7) else 'left', wrap=(j == 9))
            ws.cell(row=r, column=j).border = BOX
        r += 1
put(r, 1, 'TOTAL', True); put(r, 5, f'=SUM(E13:E{r-1})', True, al='center')
for j in range(1, 10): ws.cell(row=r, column=j).fill = PatternFill('solid', fgColor=GREY); ws.cell(row=r, column=j).border = BOX
inv_end = r

# --- avanço ---
pr = r + 2
put(pr, 1, 'AVANÇO DA EXECUÇÃO (atualiza automaticamente a partir das abas)', True, 11, NAVY)
prog_h = ['ABA', 'TOTAL DE ESTACAS', 'TERRAPLANAGEM', 'ESTAQUEAMENTO', 'PIT', 'ARRASAMENTO', '% GERAL']
for j, h in enumerate(prog_h, 1): hdr(ws.cell(row=pr + 1, column=j), h)
ws.row_dimensions[pr + 1].height = 30
PROG_FIRST = pr + 2
SHEET_ROWS = {}   # preenchido depois

# ───────────────────────────── ABAS DE DADOS ─────────────────────────────
COLS = [('A', 'Nº', 6), ('B', 'PONTO (TAG)', 14), ('C', 'ESTRUTURA', 26),
        ('D', 'E (E3D/PDMS)', 14), ('E', 'N (E3D/PDMS)', 14),
        ('F', 'E (GLOBAL) SIRGAS 2000', 17), ('G', 'N (GLOBAL) SIRGAS 2000', 17),
        ('H', 'CONF.', 9), ('I', 'Ø', 7), ('J', 'COTA DE ARRAS.', 13), ('K', 'COMPR. (m)', 10),
        ('L', 'STATUS', 15), ('M', 'DATA', 11), ('N', 'STATUS', 15), ('O', 'DATA', 11),
        ('P', 'STATUS', 15), ('Q', 'DATA', 11), ('R', 'STATUS', 15), ('S', 'DATA', 11),
        ('T', '% CONCL.', 10), ('U', 'OBSERVAÇÕES', 34)]

for aba, k in ABAS:
    v = D[k]
    s = wb.create_sheet(aba)
    s.sheet_view.showGridLines = False
    for col, _, w in COLS: s.column_dimensions[col].width = w

    s.merge_cells('A1:U1')
    put_c = s.cell(row=1, column=1, value=f'CONTROLE DE ESTAQUEAMENTO — {aba.upper()}')
    put_c.font = Font(name=F, size=14, bold=True, color='FFFFFF')
    put_c.fill = PatternFill('solid', fgColor=NAVY)
    put_c.alignment = Alignment(horizontal='center', vertical='center')
    s.row_dimensions[1].height = 28

    for rr, txt in ((2, f"Desenho: {v['doc']}  |  Rev. {v['rev']}  |  {v['data']}  |  {v['titulo']}"),
                    (3, f"Coordenadas em metro. PDMS (E3D/NAVIS) e GLOBAL SIRGAS 2000 — E(PDMS)=0 → {DATUM_E:,.3f} m ; N(PDMS)=0 → {DATUM_N:,.3f} m (ver aba Resumo)".replace(',', 'X').replace('.', ',').replace('X', '.')),
                    (4, f"Origem das coordenadas: {v['fonte']}")):
        s.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=21)
        c = s.cell(row=rr, column=1, value=txt)
        c.font = Font(name=F, size=9, bold=(rr == 2), color='1F3864' if rr == 2 else '595959')
        c.fill = PatternFill('solid', fgColor=LBLUE if rr == 2 else 'FFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center')

    H1, H2, H3 = 6, 7, 8
    for rr in (H1, H2, H3): s.row_dimensions[rr].height = 20
    s.row_dimensions[H3].height = 26
    for col in ('A', 'B', 'C', 'H', 'T', 'U'):
        s.merge_cells(f'{col}{H1}:{col}{H3}')
    s.merge_cells(f'D{H1}:G{H1}'); s.merge_cells(f'I{H1}:K{H1}'); s.merge_cells(f'L{H1}:S{H1}')
    for col in ('D', 'E', 'F', 'G', 'I', 'J', 'K'):
        s.merge_cells(f'{col}{H2}:{col}{H3}')
    for a, b in (('L', 'M'), ('N', 'O'), ('P', 'Q'), ('R', 'S')):
        s.merge_cells(f'{a}{H2}:{b}{H2}')

    hdr(s[f'A{H1}'], 'Nº'); hdr(s[f'B{H1}'], 'PONTO (TAG)'); hdr(s[f'C{H1}'], 'ESTRUTURA')
    hdr(s[f'D{H1}'], 'COORDENADAS DE LOCAÇÃO', NAVY)
    hdr(s[f'D{H2}'], 'E (E3D/PDMS)'); hdr(s[f'E{H2}'], 'N (E3D/PDMS)')
    hdr(s[f'F{H2}'], 'E (GLOBAL)\nSIRGAS 2000'); hdr(s[f'G{H2}'], 'N (GLOBAL)\nSIRGAS 2000')
    hdr(s[f'H{H1}'], 'CONF.\nDATUM')
    hdr(s[f'I{H1}'], 'DADOS DA ESTACA', NAVY)
    hdr(s[f'I{H2}'], 'Ø'); hdr(s[f'J{H2}'], 'COTA DE\nARRAS.'); hdr(s[f'K{H2}'], 'COMPR.\n(m)')
    hdr(s[f'L{H1}'], 'ESTAQUEAMENTO — CONTROLE DE EXECUÇÃO', NAVY)
    for a, nome in (('L', 'TERRAPLANAGEM'), ('N', 'ESTAQUEAMENTO'), ('P', 'PIT'), ('R', 'ARRASAMENTO')):
        hdr(s[f'{a}{H2}'], nome)
    for a in ('L', 'N', 'P', 'R'):
        hdr(s[f'{a}{H3}'], 'STATUS', '4472C4', 8)
        hdr(s[f'{chr(ord(a)+1)}{H3}'], 'DATA', '4472C4', 8)
    hdr(s[f'T{H1}'], '%\nCONCL.'); hdr(s[f'U{H1}'], 'OBSERVAÇÕES')

    row = H3 + 1
    first = row
    n = 0
    for g in v['grupos']:
        for p in g['piles']:
            n += 1
            s.cell(row=row, column=1, value=n).number_format = '0'
            s.cell(row=row, column=2, value=p['tag'])
            s.cell(row=row, column=3, value=g['estrutura'])
            s.cell(row=row, column=4, value=p['e']).number_format = '#,##0.000'
            s.cell(row=row, column=5, value=p['n']).number_format = '#,##0.000'
            s.cell(row=row, column=6, value=p['eg']).number_format = '#,##0.000'
            s.cell(row=row, column=7, value=p['ng']).number_format = '#,##0.000'
            s.cell(row=row, column=8, value=(f'=IF(AND(ABS(D{row}+Resumo!$C$6-F{row})<=0.01,'
                                             f'ABS(E{row}+Resumo!$C$7-G{row})<=0.01),"OK","VERIFICAR")'))
            s.cell(row=row, column=9, value=g['diam'])
            s.cell(row=row, column=10, value=g['arras'])
            s.cell(row=row, column=11, value=g['lest']).number_format = '#,##0.00'
            for c in (12, 14, 16, 18):
                s.cell(row=row, column=c, value='NÃO INICIADO')
                s.cell(row=row, column=c + 1).number_format = 'DD/MM/YYYY'
            s.cell(row=row, column=20, value=(f'=(COUNTIF(L{row},"CONCLUÍDO")+COUNTIF(N{row},"CONCLUÍDO")'
                                              f'+COUNTIF(P{row},"CONCLUÍDO")+COUNTIF(R{row},"CONCLUÍDO"))/4')
                   ).number_format = '0%'
            for j in range(1, 22):
                c = s.cell(row=row, column=j)
                c.font = Font(name=F, size=9, bold=(j == 2),
                              color='0000FF' if j in (12, 13, 14, 15, 16, 17, 18, 19, 21) else '000000')
                c.border = BOX
                c.alignment = Alignment(horizontal='left' if j in (3, 21) else 'center', vertical='center')
                if j in (12, 13, 14, 15, 16, 17, 18, 19, 21):
                    c.fill = PatternFill('solid', fgColor=YELLOW)
                elif n % 2 == 0:
                    c.fill = PatternFill('solid', fgColor=GREY)
            row += 1
    last = row - 1

    # linha de total
    s.cell(row=row, column=1, value='TOTAL').font = Font(name=F, size=10, bold=True)
    s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    s.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    s.cell(row=row, column=4, value=f'=COUNTA(B{first}:B{last})').font = Font(name=F, size=10, bold=True)
    s.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    s.cell(row=row, column=5, value='estacas').font = Font(name=F, size=9, italic=True)
    for a in ('L', 'N', 'P', 'R'):
        c = s[f'{a}{row}']
        c.value = f'=COUNTIF({a}{first}:{a}{last},"CONCLUÍDO")&" / "&COUNTA(B{first}:B{last})'
        c.font = Font(name=F, size=9, bold=True); c.alignment = Alignment(horizontal='center')
    s.cell(row=row, column=20, value=f'=AVERAGE(T{first}:T{last})').number_format = '0%'
    s.cell(row=row, column=20).font = Font(name=F, size=10, bold=True)
    s.cell(row=row, column=20).alignment = Alignment(horizontal='center')
    for j in range(1, 22):
        c = s.cell(row=row, column=j)
        c.fill = PatternFill('solid', fgColor='D6DCE4')
        c.border = Border(left=thin, right=thin, top=med, bottom=med)
    total_row = row

    # validação, formatação condicional, filtro, congelamento
    dv = DataValidation(type='list', formula1=STATUS, allow_blank=False, showDropDown=False)
    dv.error = 'Selecione: NÃO INICIADO, EM ANDAMENTO, CONCLUÍDO ou N/A'
    dv.errorTitle = 'Status inválido'; dv.prompt = 'Escolha o status da etapa'; dv.promptTitle = 'Status'
    s.add_data_validation(dv)
    for a in ('L', 'N', 'P', 'R'):
        dv.add(f'{a}{first}:{a}{last}')
        rng = f'{a}{first}:{a}{last}'
        s.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"CONCLUÍDO"'],
            fill=PatternFill('solid', bgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREENT)))
        s.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"EM ANDAMENTO"'],
            fill=PatternFill('solid', bgColor=AMBER), font=Font(name=F, size=9, color=AMBERT)))
        s.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"NÃO INICIADO"'],
            fill=PatternFill('solid', bgColor=REDF), font=Font(name=F, size=9, color=REDT)))
    s.conditional_formatting.add(f'H{first}:H{last}', CellIsRule(operator='equal', formula=['"VERIFICAR"'],
        fill=PatternFill('solid', bgColor=REDF), font=Font(name=F, size=9, bold=True, color=REDT)))
    s.auto_filter.ref = f'A{H3}:U{last}'
    s.freeze_panes = f'D{first}'
    s.print_title_rows = f'{H1}:{H3}'
    s.page_setup.orientation = 'landscape'
    s.page_setup.fitToWidth = 1; s.page_setup.fitToHeight = 0
    s.sheet_properties.pageSetUpPr.fitToPage = True
    SHEET_ROWS[aba] = (first, last, total_row)

# ─────────────────── RESUMO: bloco de avanço + legenda ───────────────────
r = PROG_FIRST
for aba, _ in ABAS:
    f_, l_, _t = SHEET_ROWS[aba]
    q = f"'{aba}'"
    put(r, 1, aba)
    put(r, 2, f'=COUNTA({q}!B{f_}:B{l_})', al='center')
    for j, col in zip(range(3, 7), ('L', 'N', 'P', 'R')):
        put(r, j, f'=COUNTIF({q}!{col}{f_}:{col}{l_},"CONCLUÍDO")', al='center')
    put(r, 7, f'=IFERROR(AVERAGE({q}!T{f_}:T{l_}),0)', al='center', fmt='0%')
    for j in range(1, 8): ws.cell(row=r, column=j).border = BOX
    r += 1
put(r, 1, 'TOTAL GERAL', True)
for j in range(2, 7):
    put(r, j, f'=SUM({get_column_letter(j)}{PROG_FIRST}:{get_column_letter(j)}{r-1})', True, al='center')
put(r, 7, f'=IFERROR(SUM(C{r},D{r},E{r},F{r})/(4*B{r}),0)', True, al='center', fmt='0%')
for j in range(1, 8):
    ws.cell(row=r, column=j).fill = PatternFill('solid', fgColor=GREY); ws.cell(row=r, column=j).border = BOX
prog_end = r

# legenda
lg = prog_end + 2
put(lg, 1, 'COMO PREENCHER', True, 11, NAVY)
put(lg + 1, 1, 'Células AMARELAS com texto AZUL são as únicas de preenchimento manual: os 4 pares STATUS/DATA e OBSERVAÇÕES.', fill=YELLOW)
ws.merge_cells(start_row=lg + 1, start_column=1, end_row=lg + 1, end_column=9)
put(lg + 2, 1, 'Colunas de coordenadas, Ø, cota e comprimento vêm dos desenhos — não editar. CONF. DATUM e % CONCL. são fórmulas.')
ws.merge_cells(start_row=lg + 2, start_column=1, end_row=lg + 2, end_column=9)
put(lg + 3, 1, 'STATUS (lista suspensa): NÃO INICIADO  •  EM ANDAMENTO  •  CONCLUÍDO  •  N/A     |     DATA: dd/mm/aaaa (data de conclusão da etapa).')
ws.merge_cells(start_row=lg + 3, start_column=1, end_row=lg + 3, end_column=9)
put(lg + 4, 1, 'CONF. DATUM confere E/N GLOBAL contra E/N PDMS + datum (tolerância 1 cm); "VERIFICAR" indica divergência a checar no desenho.')
ws.merge_cells(start_row=lg + 4, start_column=1, end_row=lg + 4, end_column=9)

ex = lg + 6
put(ex, 1, 'EXEMPLO DE PREENCHIMENTO DE UMA LINHA', True, 11, NAVY)
ex_h = ['PONTO', 'TERRAPLANAGEM', 'DATA', 'ESTAQUEAMENTO', 'DATA', 'PIT', 'DATA', 'ARRASAMENTO', 'DATA']
for j, h in enumerate(ex_h, 1): hdr(ws.cell(row=ex + 1, column=j), h)
ws.row_dimensions[ex + 1].height = 28
ex_v = ['ECC-001', 'CONCLUÍDO', '05/09/2026', 'CONCLUÍDO', '12/09/2026', 'EM ANDAMENTO', '', 'NÃO INICIADO', '']
for j, x in enumerate(ex_v, 1):
    put(ex + 2, j, x, al='center', color='0000FF' if j > 1 else '000000', fill=YELLOW if j > 1 else None)
    ws.cell(row=ex + 2, column=j).border = BOX

nt = ex + 4
put(nt, 1, 'NOTAS', True, 11, NAVY)
notas = [
 '1. PCE não consta desta planilha — conforme definido, as estacas do PCE serão definidas pela Engenharia.',
 '2. Torre T-5700002: o desenho DE-5400.00-5700-121-TX3-003 não traz TABELA DE COORDENADAS. As coordenadas das 96 estacas '
 '(E1-T-2 a E96-T-2) foram calculadas a partir da malha do desenho (passo de 150 cm, esc. 1:75), ancorada nas duas chamadas de '
 'coordenada da planta (E=5.006,654 / N=3.193,930 e E=5.014,154 / N=3.186,430). Recomenda-se confirmar com a Projetista antes da locação em campo.',
 '3. Reatores: nas coordenadas GLOBAIS tabeladas no desenho -006 há diferenças de até 9 mm em relação a PDMS + datum, por '
 'arredondamento do próprio desenho (que usa 714.955,43 / 7.488.954,39). Mantidos os valores originais do desenho.',
 '4. O PDF também contém o estaqueamento do PIPE-RACK (desenho DE-5400.00-5700-121-TX3-009, estacas EPR-001 a EPR-224 e PT1 a PT5), '
 'não solicitado nesta planilha e por isso não incluído — pode ser acrescentado como nova aba se necessário.',
 '5. Todas as coordenadas das abas Casa dos compressores, Fornos e Reatores foram transcritas integralmente das TABELAS DE '
 'COORDENADAS dos respectivos desenhos (216 estacas no total, conferidas uma a uma).',
]
for i, t in enumerate(notas):
    ws.merge_cells(start_row=nt + 1 + i, start_column=1, end_row=nt + 1 + i, end_column=9)
    c = put(nt + 1 + i, 1, t, size=9, wrap=True)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[nt + 1 + i].height = 26 if len(t) < 160 else 40

for col, w in zip('ABCDEFGHI', (26, 30, 22, 10, 10, 16, 17, 30, 34)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A3'

wb.save(SAIDA)
print(f'gerado: {SAIDA}\nabas:', wb.sheetnames)
print('linhas por aba:', SHEET_ROWS)
