# -*- coding: utf-8 -*-
"""Reemite a planilha de importacao de FTs com os quantitativos da memoria de calculo."""
import openpyxl, datetime as dt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "planilha.xlsx"
OUT = "/home/user/ProjectX-/engenharia/terraplanagem-u6310/IMPORTACAO_TERRAPLAN_TANQUE_Rev1-MC.xlsx"

src = openpyxl.load_workbook(SRC)
orig = src["IMPORTAÇÃO "]
HEADERS = [orig.cell(row=1, column=c).value for c in range(1, 26)]

# Quantitativos apurados na memoria de calculo (corte geometrico por linha, tabela do desenho)
CORTE_LINHA = {"816": 4100.16, "817": 4130.06, "818": 3599.36, "815": 2649.19, "824": 2789.14}
QTD_TANQUES = {"816": 4, "817": 4, "818": 4, "815": 4, "824": 1}
TANQ_FT     = {"816": 2, "817": 2, "818": 2, "815": 2, "824": 0.5}   # 824: tanque unico dividido em 2 FTs
REATERRO    = {"816": 3945.84, "817": 3947.50, "818": 3455.05, "815": 2492.22, "824": 1756.44}
BASE_ANT    = {"816": 2434.25, "817": 2464.06, "818": 2071.27, "815": 1415.86, "824": 2235.88}

FTS = [
 ("FT-CV-2619","TQ-6310816D/C",dt.date(2026,8,28),"U6310","U6310-C-610246000","816","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2620","TQ-6310816B/A",dt.date(2026,9,4), "U6310","U6310-C-610246000","816","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2621","TQ-6310817D/C",dt.date(2026,9,11),"U6310","U6310-C-610250000","817","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2622","TQ-6310817B/A",dt.date(2026,9,18),"U6310","U6310-C-610250000","817","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2623","TQ-6312824_1", dt.date(2026,9,25),"U6312","U6312-C-710073000","824","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6312"),
 ("FT-CV-2624","TQ-6312824_2", dt.date(2026,10,2),"U6312","U6312-C-710073000","824","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6312"),
 ("FT-CV-2625","TQ-6310818D/C",dt.date(2026,10,9),"U6310","U6310-C-620256000","818","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2626","TQ-6310818B/A",dt.date(2026,10,16),"U6310","U6310-C-620256000","818","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2627","TQ-6310815D/C",dt.date(2026,10,23),"U6310","U6310-C-620252000","815","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
 ("FT-CV-2628","TQ-6310815B/A",dt.date(2026,10,30),"U6310","U6310-C-620252000","815","EXECUÇÃO DA TERRAPLANAGEM NO PÁTIO DE TANCAGEM U-6310"),
]

ARIAL = Font(name="Arial", size=10)
BOLD  = Font(name="Arial", size=10, bold=True)
HDR   = Font(name="Arial", size=9, bold=True, color="FFFFFF")
SMALL = Font(name="Arial", size=8, italic=True, color="595959")
AZ    = Font(name="Arial", size=10, color="0000FF")
FH    = PatternFill("solid", fgColor="1F3864")
FP    = PatternFill("solid", fgColor="FFF2CC")
thin  = Side(style="thin", color="BFBFBF")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ---------- aba PARAMETROS ----------
pa = wb.active; pa.title = "PARÂMETROS"
pa.sheet_view.showGridLines = False
pa["A1"] = "PARÂMETROS DE CÁLCULO DA QUANTIDADE DAS FICHAS DE TAREFA"
pa["A1"].font = Font(name="Arial", size=13, bold=True, color="1F3864")
pa["A2"] = "Origem: MC-TERRAPLANAGEM_U-6310_U-6312_Rev0.xlsx — abas QUANT-LINHA e DISTRIB-FT"
pa["A2"].font = SMALL
for col, w in zip("ABCDE", [10, 46, 14, 10, 66]):
    pa.column_dimensions[col].width = w
for i, h in enumerate(["Item", "Parâmetro", "Valor", "Unid.", "Observação"], start=1):
    c = pa.cell(row=4, column=i, value=h); c.font = HDR; c.fill = FH; c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center")
PARS = [
 ("1", "Fator de empolamento (corte → solto)", 1.40, "—",
  "Fator vigente na planilha original. Confirmar por ensaio / ET-5400.00-6310-113-TX3-001 antes da medição."),
 ("2", "Cota de limpeza do terreno", 19.00, "m", "Nota da tabela de quantitativos e Nota 7 dos cortes."),
 ("3", "Cota do greide de projeto sob o tanque", 17.50, "m", "'GREIDE DE PROJETO EL.17,50' — cortes A-A a E-E."),
 ("4", "Profundidade de escavação", 1.50, "m", "Calculado: Item 2 − Item 3. Coerente com a Nota 4 do desenho."),
]
r = 5
for it, nome, val, un, obs in PARS:
    for col, v, f in [(1, it, BOLD), (2, nome, ARIAL), (3, val, AZ), (4, un, ARIAL), (5, obs, SMALL)]:
        c = pa.cell(row=r, column=col, value=v); c.font = f; c.border = BOX
        if col == 3:
            c.number_format = '#,##0.00'; c.fill = FP
            c.alignment = Alignment(horizontal="center")
        if col in (1, 4): c.alignment = Alignment(horizontal="center")
        if col == 5: c.alignment = Alignment(wrap_text=True, vertical="center")
    pa.row_dimensions[r].height = 26
    r += 1

r += 1
pa.cell(row=r, column=2, value="CORTE GEOMÉTRICO POR LINHA DE TANQUES — TABELA DO DESENHO DE-5400.00-6310-113-TX3-003 rev.0").font = BOLD
r += 1
for i, h in enumerate(["Linha", "Local", "CORTE (m³)", "REATERRO (m³)", "Nº de tanques", "Tanques por FT"], start=1):
    c = pa.cell(row=r, column=i, value=h); c.font = HDR; c.fill = FH; c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
pa.column_dimensions["F"].width = 15
hdr_lin = r; r += 1
LOCAIS = {"816": "TQ-6310816A ATÉ TQ-6310816D E BACIAS", "817": "TQ-6310817A ATÉ TQ-6310817D E BACIAS",
          "818": "TQ-6310818A ATÉ TQ-6310818D E BACIAS", "815": "TQ-6310815A ATÉ TQ-6310815D E BACIAS",
          "824": "TQ-6312824 E BACIA"}
lin_row = {}
for lin in ["816", "817", "818", "815", "824"]:
    lin_row[lin] = r
    for col, v, f in [(1, lin, BOLD), (2, LOCAIS[lin], ARIAL), (3, CORTE_LINHA[lin], AZ),
                      (4, REATERRO[lin], AZ), (5, QTD_TANQUES[lin], AZ), (6, TANQ_FT[lin], AZ)]:
        c = pa.cell(row=r, column=col, value=v); c.font = f; c.border = BOX
        if col in (3, 4): c.number_format = '#,##0.00'
        if col == 6: c.number_format = '#,##0.0'
        if col in (1, 5, 6): c.alignment = Alignment(horizontal="center")
    r += 1
r += 1
pa.cell(row=r, column=2, value="Quantidade da FT = CORTE da linha × (tanques da FT ÷ nº de tanques da linha) × fator de empolamento.").font = SMALL
r += 1
pa.cell(row=r, column=2, value="Para o TQ-6312824 — tanque único dividido em 2 FTs — 'Tanques por FT' vale 0,5, produzindo a mesma fração de 1/2.").font = SMALL
r += 2
pa.cell(row=r, column=2, value="TOTAL DAS 10 FICHAS DE TAREFA (m³, volume solto)").font = BOLD
tc = pa.cell(row=r, column=3, value="=SUM('IMPORTAÇÃO '!I2:I11)")
tc.font = BOLD; tc.number_format = '#,##0.00'
tc.fill = PatternFill("solid", fgColor="DDEBF7"); tc.border = BOX
pa.cell(row=r, column=4, value="m³").font = ARIAL
pa.cell(row=r, column=5, value="Confere com o item 2 do RESUMO da memória de cálculo (24.175,07 m³).").font = SMALL
r += 1
pa.cell(row=r, column=2, value="A aba IMPORTAÇÃO não traz linha de total, para que uma rotina que varra até a última linha preenchida não a leia como registro.").font = SMALL

# ---------- aba IMPORTACAO ----------
ws = wb.create_sheet("IMPORTAÇÃO ")   # espaco final preservado do arquivo original
ws.sheet_view.showGridLines = False
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=i, value=h); c.font = HDR; c.fill = FH; c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 34
ws.freeze_panes = "A2"

OBS = ("Quantidade revisada conforme memória de cálculo MC-TERRAPLANAGEM_U-6310_U-6312_Rev0, "
       "com base no DE-5400.00-6310-113-TX3-003 rev.0. O ID usado sofrerá alterações com a aprovação do Book B.")
ATIV = "001-CORTE;002-TROCA DE SOLO;003-ATERRO;004-COMPACTAÇÃO"

r = 2
for chave, tag, data, unid, prim, lin, escopo in FTS:
    lr = lin_row[lin]
    vals = {
      1: tag, 2: "TERRAPLANAGEM DE IMPLANTAÇÃO", 3: "CIVIL", 4: "N/A", 5: chave, 6: data,
      7: unid, 8: prim,
      9: f"=PARÂMETROS!$C${lr}*(PARÂMETROS!$F${lr}/PARÂMETROS!$E${lr})*PARÂMETROS!$C$5",
      10: "M³", 11: OBS, 12: "brener.cabral", 13: escopo,
      14: "DE-5400.00-6310-113-TX3-003", 15: "EXEC", 18: "PRODUCAO", 19: ATIV, 25: "N/A",
    }
    for col in range(1, 26):
        c = ws.cell(row=r, column=col, value=vals.get(col))
        c.font = ARIAL; c.border = BOX
        if col == 6: c.number_format = 'dd/mm/yyyy'; c.alignment = Alignment(horizontal="center")
        if col == 9: c.number_format = '#,##0.00'; c.font = BOLD
    r += 1
last = r - 1
# Nenhuma linha de total e escrita nesta aba: uma rotina de importacao que varre ate a
# ultima linha preenchida leria o total como um registro. O total fica na aba PARAMETROS.

widths = [17, 30, 9, 7, 13, 14, 9, 20, 13, 7, 60, 15, 48, 30, 11, 9, 10, 13, 46, 13, 13, 13, 16, 13, 11]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- aba CONTROLE ----------
cs = wb.create_sheet("CONTROLE DE REVISÃO")
cs.sheet_view.showGridLines = False
cs["A1"] = "CONTROLE DE REVISÃO — QUANTIDADES DAS FICHAS DE TAREFA"
cs["A1"].font = Font(name="Arial", size=13, bold=True, color="1F3864")
for col, w in zip("ABCDEF", [16, 20, 18, 18, 16, 12]):
    cs.column_dimensions[col].width = w
for i, h in enumerate(["Chave de importação", "TAG", "Quantidade anterior (m³)",
                       "Quantidade revisada (m³)", "Diferença (m³)", "Variação"], start=1):
    c = cs.cell(row=3, column=i, value=h); c.font = HDR; c.fill = FH; c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cs.row_dimensions[3].height = 34
r = 4
for i, (chave, tag, data, unid, prim, lin, escopo) in enumerate(FTS):
    ant = BASE_ANT[lin] / 2 * 0.7 * 1.4
    for col, v, f, fmt in [(1, chave, BOLD, None), (2, tag, ARIAL, None), (3, ant, AZ, '#,##0.00'),
                           (4, f"='IMPORTAÇÃO '!I{i+2}", ARIAL, '#,##0.00'),
                           (5, f"=D{r}-C{r}", ARIAL, '#,##0.00'),
                           (6, f"=D{r}/C{r}-1", ARIAL, '0.0%')]:
        c = cs.cell(row=r, column=col, value=v); c.font = f; c.border = BOX
        if fmt: c.number_format = fmt
    r += 1
cs.cell(row=r, column=2, value="TOTAL").font = BOLD
for col, f in [(3, f"=SUM(C4:C{r-1})"), (4, f"=SUM(D4:D{r-1})"), (5, f"=D{r}-C{r}"), (6, f"=D{r}/C{r}-1")]:
    c = cs.cell(row=r, column=col, value=f); c.font = BOLD; c.border = BOX
    c.number_format = '0.0%' if col == 6 else '#,##0.00'
    c.fill = PatternFill("solid", fgColor="DDEBF7")
r += 2
for t in ["Quantidade anterior: fórmula da planilha original =(BASE/2)×0,7×1,4, com bases 2.434,25 / 2.464,06 / 2.071,27 / 1.415,86 / 2.235,88 m³.",
          "Essas bases não são rastreáveis aos desenhos emitidos em 07/08/2026 e o fator 0,7 não corresponde à geometria do projeto",
          "(a escavação sob os tanques responde por 94% a 96% do corte nas linhas de U-6310). Ver aba CONFRONTO-FT da memória de cálculo."]:
    cs.cell(row=r, column=1, value=t).font = SMALL; r += 1

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("gerado:", OUT)
